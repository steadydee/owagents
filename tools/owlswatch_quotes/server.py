#!/usr/bin/env python3
"""Narrow MCP tools for Owl's Watch quote drafting.

This server intentionally exposes only clerk actions. It reads credentials from
environment/config, never from tool parameters, and always returns structured
JSON-compatible results.
"""

from __future__ import annotations

import datetime as dt
import base64
import hashlib
import html
import json
import os
import re
import sys
import tempfile
import traceback
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Callable

WORKSPACE = Path(os.environ.get("OWLSWATCH_COTIZA_WORKSPACE", "~/.openclaw/workspace-owlswatch-cotiza")).expanduser()
CONFIG_PATH = Path(os.environ.get("OPENCLAW_CONFIG_PATH", "~/.openclaw-owlswatch/openclaw.json")).expanduser()
MEMORY_DIR = WORKSPACE / "memory"
MOCK_QUOTES_DIR = WORKSPACE / "mock" / "quotes"
MOCK_DRIVE_DIR = WORKSPACE / "mock" / "drive"
QUOTE_LOGO_PATH = WORKSPACE / "assets" / "WATERMARK FULL LOGO.png"
DEFAULT_API_BASE_URL = "https://operations.owlswatch.com"
QUOTE_RULE_VERSION = "2027-one-page-quote-layout-v1"
QUOTE_RATES = {
    2026: {
        "pricebook_version": "2026-operators",
        "cabin_rack": 840000,
        "cabin_operator_net": 756000,
        "extra_person": 110000,
        "guide_room": 316000,
        "bird_tour_rack": 150000,
        "bird_tour_operator_net": None,
        "afternoon_extension": 100000,
        "bilingual_half_day": 200000,
        "bilingual_full_day": 300000,
        "breakfast": 55000,
        "lunch": 65000,
        "dinner": 65000,
        "guide_driver_breakfast": 0,
        "guide_driver_lunch": 25000,
        "guide_driver_dinner": 25000,
    },
    2027: {
        "pricebook_version": "2027-operators",
        "cabin_rack": 880000,
        "cabin_operator_net": 792000,
        "extra_person": 120000,
        "guide_room": 350000,
        "bird_tour_rack": 160000,
        "bird_tour_operator_net": 144000,
        "afternoon_extension": 110000,
        "bilingual_half_day": 220000,
        "bilingual_full_day": 320000,
        "breakfast": 60000,
        "lunch": 70000,
        "dinner": 70000,
        "guide_driver_breakfast": 0,
        "guide_driver_lunch": 35000,
        "guide_driver_dinner": 35000,
    },
}

SAFE_ID_RE = re.compile(r"^[A-Za-z0-9_.:@/+\\-]{1,220}$")
TEXT_RE = re.compile(r"^[\s\S]{0,20000}$")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


class ToolError(Exception):
    def __init__(self, code: str, message: str, retryable: bool = False):
        super().__init__(message)
        self.code = code
        self.message = message
        self.retryable = retryable


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat().replace("+00:00", "Z")


def local_now_label() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def sanitized_error(exc: Exception) -> dict[str, Any]:
    if isinstance(exc, ToolError):
        return {"ok": False, "error": {"code": exc.code, "message": exc.message, "retryable": exc.retryable}}
    return {"ok": False, "error": {"code": "internal_error", "message": "Tool failed without exposing sensitive details.", "retryable": False}}


def load_config() -> dict[str, Any]:
    try:
        return json.loads(CONFIG_PATH.read_text())
    except FileNotFoundError:
        return {}
    except json.JSONDecodeError as exc:
        raise ToolError("config_invalid", f"OpenClaw config is not valid JSON: line {exc.lineno}") from exc


def cfg_env(config: dict[str, Any], key: str) -> str | None:
    value = os.environ.get(key)
    if value:
        return value
    mcp_env = config.get("mcp", {}).get("servers", {}).get("owlswatch_quotes", {}).get("env", {})
    value = mcp_env.get(key)
    if isinstance(value, str) and value and not value.startswith("<"):
        return value
    env_vars = config.get("env", {}).get("vars", {})
    value = env_vars.get(key)
    if isinstance(value, str) and value and not value.startswith("<"):
        return value
    quotes_cfg = config.get("quotes", {})
    camel_key = key.lower().split("_")
    candidate = camel_key[0] + "".join(part.title() for part in camel_key[1:])
    value = quotes_cfg.get(candidate)
    if isinstance(value, str) and value and not value.startswith("<"):
        return value
    return None


def mocks_enabled(config: dict[str, Any]) -> bool:
    value = cfg_env(config, "OWLSWATCH_QUOTES_MOCKS")
    return str(value).lower() in {"1", "true", "yes", "on"}


def operations_base_url(config: dict[str, Any]) -> str:
    raw = cfg_env(config, "OPERATIONS_BASE_URL") or cfg_env(config, "OPERATIONS_API_BASE_URL") or DEFAULT_API_BASE_URL
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme != "https" or not parsed.netloc:
        raise ToolError("config_invalid", "Operations base URL must be an https URL.")
    return raw.rstrip("/")


def operations_token(config: dict[str, Any]) -> str:
    token = cfg_env(config, "QUOTE_INTAKE_API_TOKEN")
    if token:
        return token
    raise ToolError("config_missing", "Operations quote intake token is missing from tool environment/config.")


def google_folder_id(config: dict[str, Any]) -> str:
    raw = cfg_env(config, "GOOGLE_DRIVE_QUOTES_FOLDER_ID")
    if not raw:
        raw = config.get("googleDrive", {}).get("quotesFolderId")
    if not isinstance(raw, str) or not raw or raw.startswith("<"):
        raise ToolError("config_missing", "Google Drive quotes folder id is missing from tool environment/config.")
    return raw


def google_template_id(config: dict[str, Any]) -> str | None:
    raw = cfg_env(config, "GOOGLE_DRIVE_QUOTE_TEMPLATE_SPREADSHEET_ID")
    if not raw:
        raw = config.get("googleDrive", {}).get("quoteTemplateSpreadsheetId")
    if isinstance(raw, str) and raw and not raw.startswith("<"):
        return raw
    return None


def google_credentials_path(config: dict[str, Any]) -> str:
    raw = cfg_env(config, "GOOGLE_APPLICATION_CREDENTIALS")
    if not raw:
        raw = config.get("googleDrive", {}).get("serviceAccountCredentials")
    if not isinstance(raw, str) or not raw or raw.startswith("<"):
        raise ToolError("config_missing", "Google service account credentials are missing from tool environment/config.")
    path = Path(raw).expanduser()
    if not path.is_file():
        raise ToolError("config_missing", "Google service account credentials file was not found.")
    return str(path)


def gmail_account(config: dict[str, Any]) -> str:
    raw = cfg_env(config, "OWLSWATCH_GMAIL_ACCOUNT")
    if not raw:
        raw = config.get("gmail", {}).get("account")
    if not isinstance(raw, str) or "@" not in raw:
        raise ToolError("config_missing", "Owl's Watch Gmail account is missing from tool environment/config.")
    return raw


def google_build_service(config: dict[str, Any], api: str, version: str, scopes: list[str], delegated_subject: str | None = None) -> Any:
    credentials_path = google_credentials_path(config)
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise ToolError("dependency_missing", "Google API Python packages are not installed in the tool runtime.") from exc
    credentials = service_account.Credentials.from_service_account_file(credentials_path, scopes=scopes)
    if delegated_subject:
        credentials = credentials.with_subject(delegated_subject)
    return build(api, version, credentials=credentials)


def validate_safe_id(name: str, value: Any) -> str:
    if not isinstance(value, (str, int)):
        raise ToolError("invalid_input", f"{name} must be a string or integer.")
    text = str(value)
    if not SAFE_ID_RE.match(text):
        raise ToolError("invalid_input", f"{name} is malformed.")
    return text


def validate_text(name: str, value: Any, required: bool = False, max_len: int = 20000) -> str | None:
    if value is None:
        if required:
            raise ToolError("invalid_input", f"{name} is required.")
        return None
    if not isinstance(value, str) or len(value) > max_len or not TEXT_RE.match(value):
        raise ToolError("invalid_input", f"{name} is malformed.")
    return value


def validate_object(name: str, value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ToolError("invalid_input", f"{name} must be an object.")
    return value


def validate_string_list(name: str, value: Any, max_items: int = 50) -> list[str]:
    if value is None:
        return []
    if not isinstance(value, list) or len(value) > max_items:
        raise ToolError("invalid_input", f"{name} must be a list.")
    result = []
    for item in value:
        text = validate_text(name, item, required=True, max_len=2000)
        result.append(text or "")
    return result


def validate_no_extra(args: dict[str, Any], allowed: set[str]) -> None:
    extra = set(args) - allowed
    if extra:
        raise ToolError("invalid_input", f"Unexpected parameter: {sorted(extra)[0]}.")


def safe_http_error_details(raw: str) -> dict[str, Any] | None:
    if not raw:
        return None
    text = raw[:2000]
    try:
        data = json.loads(text)
    except Exception:
        return {"body": re.sub(r"\s+", " ", text).strip()[:800]} if text.strip() else None
    if not isinstance(data, dict):
        return {"body": str(data)[:800]}
    allowed: dict[str, Any] = {}
    for key in ("error", "message", "code", "details", "issues", "fields"):
        if key in data:
            allowed[key] = data[key]
    return allowed or None


def http_json(method: str, url: str, payload: dict[str, Any], headers: dict[str, str], timeout: int = 45, attempts: int = 1) -> dict[str, Any]:
    body = json_dumps(payload).encode()
    last_error: ToolError | None = None
    for attempt in range(max(1, attempts)):
        req = urllib.request.Request(url, data=body, headers={**headers, "Content-Type": "application/json"}, method=method)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode()
                return json.loads(raw or "{}")
        except urllib.error.HTTPError as exc:
            retryable = 500 <= exc.code <= 599 or exc.code in (408, 409, 425, 429)
            raw = exc.read().decode(errors="replace")
            details = safe_http_error_details(raw)
            suffix = f" Details: {json_dumps(details)}" if details else ""
            last_error = ToolError("http_error", f"Upstream request failed with HTTP {exc.code}.{suffix}", retryable=retryable)
            if not retryable or attempt + 1 >= attempts:
                raise last_error from exc
        except urllib.error.URLError as exc:
            last_error = ToolError("network_error", "Upstream request failed due to a network error.", retryable=True)
            if attempt + 1 >= attempts:
                raise last_error from exc
    raise last_error or ToolError("network_error", "Upstream request failed.", retryable=True)


def http_get_json(url: str, headers: dict[str, str], timeout: int = 45) -> dict[str, Any]:
    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode()
            return json.loads(raw or "{}")
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode(errors="replace")
        details = safe_http_error_details(raw)
        suffix = f" Details: {json_dumps(details)}" if details else ""
        raise ToolError("http_error", f"Upstream request failed with HTTP {exc.code}.{suffix}", retryable=500 <= exc.code <= 599) from exc
    except urllib.error.URLError as exc:
        raise ToolError("network_error", "Upstream request failed due to a network error.", retryable=True) from exc


def bearer_headers(config: dict[str, Any]) -> dict[str, str]:
    return {"Authorization": f"Bearer {operations_token(config)}"}


def coerce_int(value: Any, default: int | None = None) -> int | None:
    if value is None:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return default


def request_summary(payload: dict[str, Any]) -> str:
    for key in ("requestSummary", "summary", "rawRequestSummary", "sourceText", "requestText", "rawRequest", "raw_request", "rawText"):
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    parsed = payload.get("parsedRequest")
    if isinstance(parsed, dict):
        for key in ("requestSummary", "summary", "rawRequestSummary"):
            value = parsed.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    source = payload.get("source")
    if isinstance(source, dict):
        value = source.get("bodyText") or source.get("text")
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def first_value(payload: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in payload:
            return payload[key]
    parsed = payload.get("parsedRequest")
    if isinstance(parsed, dict):
        for key in keys:
            if key in parsed:
                return parsed[key]
    return None


def object_value(payload: dict[str, Any], key: str) -> dict[str, Any]:
    value = payload.get(key)
    return value if isinstance(value, dict) else {}


def string_value(value: Any) -> str | None:
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on", "redo", "recreate", "regenerate"}
    return False


def date_value(value: Any) -> str | None:
    text = string_value(value)
    if not text:
        return None
    normalized = text.replace("/", "-")
    return normalized if DATE_RE.match(normalized) else text


def year_from_iso_date(value: Any) -> int | None:
    text = date_value(value)
    if not text or not DATE_RE.match(text):
        return None
    return int(text[:4])


def rate_year_from_payload(payload: dict[str, Any], calc: dict[str, Any] | None = None) -> int:
    explicit = coerce_int(first_value(payload, "year", "rateYear", "rate_year"), None)
    if explicit:
        return explicit
    arrival_year = year_from_iso_date(first_value(payload, "arrivalDate", "arrival_date", "arrival", "startDate", "start_date", "date", "visitDate", "visit_date", "checkIn", "check_in"))
    if arrival_year:
        return arrival_year
    departure_year = year_from_iso_date(first_value(payload, "departureDate", "departure_date", "departure", "endDate", "end_date", "date", "visitDate", "visit_date", "checkOut", "check_out"))
    if departure_year:
        return departure_year
    version = string_value(calc.get("pricebookVersion")) if isinstance(calc, dict) else None
    if version:
        match = re.search(r"\b(20\d{2})\b", version)
        if match:
            return int(match.group(1))
    return 2026


def quote_rates_for_year(year: int) -> dict[str, Any]:
    return QUOTE_RATES.get(year, QUOTE_RATES[2026])


def infer_audience(payload: dict[str, Any], summary: str) -> str | None:
    for key in (
        "audience",
        "quoteAudience",
        "quoteType",
        "quote_type",
        "operatorOrDirect",
        "operator_or_direct",
        "customerType",
        "customer_type",
        "clientType",
        "client_type",
        "rateType",
        "rate_type",
    ):
        value = first_value(payload, key)
        if isinstance(value, str):
            lowered = value.strip().lower()
            if lowered in {"operator", "agency", "agent", "b2b", "net"}:
                return "operator"
            if lowered in {"direct", "retail", "rack", "client", "guest"}:
                return "direct"
    for key in ("isOperator", "is_operator", "operator"):
        value = first_value(payload, key)
        if value is True:
            return "operator"
        if value is False and key in {"isOperator", "is_operator"}:
            return "direct"
    lowered_summary = summary.lower()
    if any(token in lowered_summary for token in ("operator", "agency", "agencia", "net rate", "tarifa operador")):
        return "operator"
    if any(token in lowered_summary for token in ("direct client", "directo", "rack rate", "retail")):
        return "direct"
    return None


def infer_guest_count(payload: dict[str, Any], summary: str) -> int | None:
    guests_payload = object_value(payload, "guests")
    adults = coerce_int(guests_payload.get("adults"), None)
    children = coerce_int(guests_payload.get("children"), 0) or 0
    if adults is not None:
        return adults + children
    direct = coerce_int(first_value(payload, "guestCount", "guest_count", "pax", "adults"), None)
    if direct is not None:
        return direct
    guests_value = payload.get("guests")
    direct = coerce_int(guests_value, None)
    if direct is not None:
        return direct
    match = re.search(r"\b(\d{1,2})\s*(?:pax|guests?|clients?|people|persons?|personas?|adults?|adultos?)\b", summary, re.I)
    if match:
        return int(match.group(1))
    if re.search(r"\b(couple|pareja)\b", summary, re.I):
        return 2
    return None


def infer_staff_count(payload: dict[str, Any], role: str) -> int:
    summary = request_summary(payload)
    keys = {
        "guide": ("outsideGuideCount", "outside_guide_count", "operatorGuideCount", "operator_guide_count", "guideCount", "guide_count"),
        "driver": ("driverCount", "driver_count", "operatorDriverCount", "operator_driver_count", "outsideDriverCount", "outside_driver_count"),
    }[role]
    for key in keys:
        value = first_value(payload, key)
        count = coerce_int(value, None)
        if count is not None:
            return max(0, count)

    staff = object_value(payload, "staff") or object_value(payload, "outsideStaff") or object_value(payload, "outside_staff")
    if staff:
        nested_keys = ("guides", "guideCount", "guide_count") if role == "guide" else ("drivers", "driverCount", "driver_count")
        for key in nested_keys:
            count = coerce_int(staff.get(key), None)
            if count is not None:
                return max(0, count)

    guests = object_value(payload, "guests")
    if role == "guide":
        count = coerce_int(guests.get("guides"), None)
        if count is not None:
            return max(0, count)
        birding = object_value(payload, "birding")
        for key in ("guides", "guideCount", "guide_count", "localGuideCount", "local_guide_count"):
            count = coerce_int(birding.get(key), None)
            if count is not None:
                return max(0, count)
        for item in item_codes(payload):
            code = item_code_text(item)
            if "guide" in code or "guia" in code or "guía" in code:
                count = coerce_int(item.get("quantity"), None)
                if count is None:
                    count = coerce_int(item.get("count"), None)
                if count is not None:
                    return max(0, count)
    else:
        count = coerce_int(guests.get("drivers"), None)
        if count is not None:
            return max(0, count)

    lowered = summary.lower()
    role_words = r"(?:guides?|gu[ií]as?)" if role == "guide" else r"(?:drivers?|conductores?|chofer(?:es)?)"
    match = re.search(rf"\b(\d{{1,2}})\s+{role_words}\b", lowered, re.I)
    if match:
        return int(match.group(1))
    match = re.search(rf"\b{role_words}\s*[:x]?\s*(\d{{1,2}})\b", lowered, re.I)
    if match:
        return int(match.group(1))

    if role == "driver" and re.search(r"\b(driver|conductor|chofer)\b", lowered, re.I):
        return 1

    external_guide = re.search(r"\b(own|outside|operator|agency|tour leader|acompa[ñn]ante)\s+(guide|gu[ií]a)\b", lowered, re.I)
    if role == "guide" and external_guide:
        return 1
    if role == "guide" and re.search(r"\b(guide and driver|gu[ií]a y (?:conductor|chofer))\b", lowered, re.I):
        return 1
    return 0


def infer_trip_leader_count(payload: dict[str, Any]) -> int:
    summary = request_summary(payload)
    for key in ("tripLeaderCount", "trip_leader_count", "tourLeaderCount", "tour_leader_count"):
        count = coerce_int(first_value(payload, key), None)
        if count is not None:
            return max(0, count)
    guests = object_value(payload, "guests")
    for key in ("tripLeaders", "trip_leaders", "tourLeaders", "tour_leaders", "leaders"):
        count = coerce_int(guests.get(key), None)
        if count is not None:
            return max(0, count)
    match = re.search(r"\b(\d{1,2})\s*(?:trip|tour)\s+leaders?\b", summary, re.I)
    if match:
        return int(match.group(1))
    match = re.search(r"\b(\d{1,2})\s*(?:acompa[ñn]antes?)\b", summary, re.I)
    if match:
        return int(match.group(1))
    return 0


def infer_staff_meal_days(payload: dict[str, Any]) -> int:
    summary = request_summary(payload)
    meals = object_value(payload, "meals")
    days = coerce_int(first_value(payload, "staffMealDays", "staff_meal_days", "tourDays", "tour_days", "birdingDays", "birding_days"), None)
    if days is None:
        days = coerce_int(meals.get("staffLunches"), None) or coerce_int(meals.get("staff_lunches"), None)
    if days is None:
        days = coerce_int(meals.get("lunches"), None)
    if days is None:
        days = coerce_int(first_value(payload, "nights", "nightCount", "night_count"), None)
    if days is None:
        match = re.search(r"\b(\d{1,2})\s*(?:days?|d[ií]as|tour days?|birding days?)\b", summary, re.I)
        if match:
            days = int(match.group(1))
    return max(1, days or 1)


def requested_services_meals(payload: dict[str, Any]) -> dict[str, Any]:
    requested = object_value(payload, "requestedServices") or object_value(payload, "requested_services")
    meals = object_value(requested, "meals")
    if meals:
        return meals
    return object_value(payload, "meals")


def staff_role_counts(payload: dict[str, Any]) -> tuple[int, int]:
    return infer_staff_count(payload, "guide"), infer_staff_count(payload, "driver")


def staff_role_label(payload: dict[str, Any]) -> str:
    guides, drivers = staff_role_counts(payload)
    if guides > 0 and drivers > 0:
        return "Guide/Driver"
    if drivers > 0:
        return "Driver"
    return "Guide"


def has_guide_room_item(items: list[dict[str, Any]]) -> bool:
    return any("guide room" in line_text(item) or "habitaci" in line_text(item) and "gu" in line_text(item) for item in items)


def requests_guide_room(payload: dict[str, Any], summary: str) -> bool:
    lodging = object_value(payload, "lodging")
    if coerce_int(first_value(payload, "guideRoomCount", "guide_room_count"), None):
        return True
    if coerce_int(lodging.get("guideRoomCount"), None) or coerce_int(lodging.get("guide_room_count"), None):
        return True
    return bool(re.search(r"\bguide room\b|\bhabitaci[oó]n\b.{0,40}\bgu[ií]a\b|\bgu[ií]a\b.{0,40}\bhabitaci[oó]n\b", summary, re.I))


def lodging_breakfast_days(payload: dict[str, Any], summary: str) -> int:
    arrival = date_only(payload.get("arrivalDate"))
    departure = date_only(payload.get("departureDate"))
    nights = infer_nights(payload, arrival, departure, summary)
    return max(0, nights or 0)


def normalized_meal_day_count(payload: dict[str, Any], meal: str) -> int | None:
    meals = requested_services_meals(payload)
    for key in (f"{meal}s", meal):
        value = meals.get(key)
        count = coerce_int(value, None)
        if count is not None:
            return max(0, count)
    return None


def staff_meal_day_count(payload: dict[str, Any], meal: str, items: list[dict[str, Any]]) -> int:
    summary = request_summary(payload)
    normalized_count = normalized_meal_day_count(payload, meal)
    if normalized_count is not None:
        if meal == "breakfast" and normalized_count == 0:
            return 0
        return normalized_count

    if meal == "breakfast":
        meals = requested_services_meals(payload)
        value = meals.get("breakfasts") or meals.get("breakfast")
        if value == "included_with_lodging":
            return lodging_breakfast_days(payload, summary) if (requests_breakfast(payload, summary) or has_guide_room_item(items) or requests_guide_room(payload, summary)) else 0
        if requests_breakfast(payload, summary):
            days = lodging_breakfast_days(payload, summary)
            return days or 1
        return 0

    if meal == "lunch":
        if requests_lunch(payload, summary):
            return infer_staff_meal_days(payload)
        return 0

    if meal == "dinner":
        if requests_dinner(payload, summary):
            days = lodging_breakfast_days(payload, summary)
            return days or infer_staff_meal_days(payload)
        return 0

    return 0


def item_codes(payload: dict[str, Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for key in ("items", "lineItems", "services", "activities", "guidedTours", "birdingTours"):
        value = payload.get(key)
        if isinstance(value, list):
            result.extend(item for item in value if isinstance(item, dict))
        elif isinstance(value, dict):
            for name, nested in value.items():
                if isinstance(nested, dict):
                    result.append({"service": name, **nested})
    return result


def item_code_text(item: dict[str, Any]) -> str:
    return " ".join(str(item.get(key) or "") for key in ("code", "serviceCode", "name", "description", "service", "type", "tourType", "schedule", "guideType")).lower()


def infer_nights(payload: dict[str, Any], arrival: str | None, departure: str | None, summary: str) -> int | None:
    nights = coerce_int(first_value(payload, "nights", "nightCount", "night_count"), None)
    if nights is not None:
        return nights
    if arrival and departure and DATE_RE.match(arrival) and DATE_RE.match(departure):
        try:
            start = dt.date.fromisoformat(arrival)
            end = dt.date.fromisoformat(departure)
            if end > start:
                return (end - start).days
        except ValueError:
            pass
    match = re.search(r"\b(\d{1,2})\s*(?:nights?|noches?)\b", summary, re.I)
    if match:
        return int(match.group(1))
    if re.search(r"\b(two nights?|dos noches)\b", summary, re.I):
        return 2
    return None


def infer_cabin_count(payload: dict[str, Any], summary: str) -> int:
    lodging_payload = object_value(payload, "lodging")
    count = coerce_int(lodging_payload.get("cabinCount"), None)
    if count is None:
        count = coerce_int(first_value(payload, "cabinCount", "cabin_count", "cabins", "rooms", "roomCount", "room_count"), None)
    if count is None:
        for item in item_codes(payload):
            if "cabin" in item_code_text(item) or "caba" in item_code_text(item):
                count = coerce_int(item.get("quantity"), 1)
                break
    if count is None:
        match = re.search(r"\b(\d{1,2})\s*(?:cabins?|caba[ñn]as?|rooms?|habitaciones?)\b", summary, re.I)
        if match:
            count = int(match.group(1))
    if count is None and re.search(r"\b(cabin|cabaña|cabana)\b", summary, re.I):
        count = 1
    return max(1, count or 1)


def has_full_board(payload: dict[str, Any], summary: str) -> bool:
    meals_payload = object_value(payload, "meals")
    if meals_payload.get("fullBoard") is True or meals_payload.get("full_board") is True:
        return True
    if first_value(payload, "fullBoard", "full_board", "foodIncluded", "food_included", "mealsIncluded", "meals_included") is True:
        return True
    meal_plan = string_value(first_value(payload, "mealPlan", "meal_plan", "meals")) or string_value(meals_payload.get("plan"))
    if meal_plan and meal_plan.lower().replace("-", "_") in {"full_board", "food_included", "meals_included"}:
        return True
    if any("full_board" in item_code_text(item) or "full board" in item_code_text(item) for item in item_codes(payload)):
        return True
    return bool(re.search(r"\b(full board|food included|meals included|comida incluida|pensi[oó]n completa|a?alimentaci[oó]n completa)\b", summary, re.I))


def meal_plan_tokens(meals_payload: dict[str, Any]) -> str:
    values = [
        meals_payload.get("plan"),
        meals_payload.get("mealPlan"),
        meals_payload.get("meal_plan"),
    ]
    return " ".join(str(value or "") for value in values).lower().replace("-", "_")


def nested_meal_count(meals_payload: dict[str, Any], meal: str) -> int | None:
    plural = f"{meal}s"
    for key in (plural, meal):
        value = meals_payload.get(key)
        if isinstance(value, dict):
            for nested_key in ("clients", "guests", "adults", "count", "quantity"):
                count = coerce_int(value.get(nested_key), None)
                if count is not None:
                    return count
        count = coerce_int(value, None)
        if count is not None:
            return count

    for key in ("guestMealCounts", "guest_meal_counts", "clientMealCounts", "client_meal_counts"):
        counts = meals_payload.get(key)
        if isinstance(counts, dict):
            count = coerce_int(counts.get(meal), None)
            if count is None:
                count = coerce_int(counts.get(plural), None)
            if count is not None:
                return count
    return None


def item_meal_count(payload: dict[str, Any], meal: str) -> int | None:
    for item in item_codes(payload):
        code = item_code_text(item)
        if meal in code or (meal == "breakfast" and "desayuno" in code) or (meal == "lunch" and "almuerzo" in code):
            count = coerce_int(item.get("quantity"), None)
            if count is None:
                count = coerce_int(item.get("count"), None)
            if count is not None:
                return count
    return None


def normalize_day_trip_meal_count(count: int | None, guest_count: int | None, staff_count: int = 0) -> int | None:
    if count is None or count <= 1:
        return count
    if guest_count and (count == guest_count or count == guest_count + staff_count):
        return 1
    if guest_count and count > guest_count:
        return max(1, round(count / guest_count))
    return count


def meal_value_requested(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, dict):
        return any(meal_value_requested(nested) for nested in value.values())
    text = string_value(value)
    if not text:
        return False
    return text.lower().strip() not in {"0", "no", "none", "excluded", "false", "included_with_lodging"}


def requests_breakfast(payload: dict[str, Any], summary: str) -> bool:
    meals_payload = object_value(payload, "meals")
    if nested_meal_count(meals_payload, "breakfast"):
        return True
    if item_meal_count(payload, "breakfast"):
        return True
    if meal_value_requested(meals_payload.get("breakfast")):
        return True
    if first_value(payload, "includeBreakfast", "include_breakfast", "breakfastIncluded", "breakfast_included") is True:
        return True
    plan = meal_plan_tokens(meals_payload)
    if "breakfast" in plan or "desayuno" in plan:
        return True
    value = string_value(first_value(payload, "breakfast", "breakfasts", "breakfastCount", "breakfast_count"))
    if value and value.lower() not in {"no", "none", "excluded", "included_with_lodging"}:
        return True
    return bool(re.search(r"\b(breakfast|desayuno)\b", summary, re.I))


def requests_lunch(payload: dict[str, Any], summary: str) -> bool:
    meals_payload = object_value(payload, "meals")
    if nested_meal_count(meals_payload, "lunch"):
        return True
    if item_meal_count(payload, "lunch"):
        return True
    if meal_value_requested(meals_payload.get("lunch")):
        return True
    if first_value(payload, "includeLunch", "include_lunch", "lunchIncluded", "lunch_included") is True:
        return True
    plan = meal_plan_tokens(meals_payload)
    if "lunch" in plan or "almuerzo" in plan:
        return True
    value = string_value(first_value(payload, "lunch", "lunches", "lunchCount", "lunch_count"))
    if value and value.lower() not in {"no", "none", "excluded"}:
        return True
    return bool(re.search(r"\b(lunch|almuerzo)\b", summary, re.I))


def requests_checkout_lunch(summary: str) -> bool:
    return bool(
        re.search(
            r"\b(?:checkout|check[ -]?out|departure|salida)\b.{0,40}\b(?:lunch|almuerzo)\b|\b(?:lunch|almuerzo)\b.{0,40}\b(?:checkout|check[ -]?out|departure|salida)\b",
            summary,
            re.I,
        )
    )


def default_cabin_lunch_count(nights: int | None, summary: str) -> int:
    if not nights:
        return 0
    if requests_checkout_lunch(summary):
        return max(0, nights)
    return max(0, nights - 1)


def requests_dinner(payload: dict[str, Any], summary: str) -> bool:
    meals_payload = object_value(payload, "meals")
    if nested_meal_count(meals_payload, "dinner"):
        return True
    if item_meal_count(payload, "dinner"):
        return True
    if meal_value_requested(meals_payload.get("dinner")):
        return True
    if first_value(payload, "includeDinner", "include_dinner", "dinnerIncluded", "dinner_included") is True:
        return True
    plan = meal_plan_tokens(meals_payload)
    if "dinner" in plan or "cena" in plan or "comida" in plan:
        return True
    value = string_value(first_value(payload, "dinner", "dinners", "dinnerCount", "dinner_count"))
    if value and value.lower() not in {"no", "none", "excluded"}:
        return True
    return bool(re.search(r"\b(dinners?|cenas?|comidas?)\b", summary, re.I))


def is_day_trip_quote(payload: dict[str, Any], arrival: str | None, departure: str | None, nights: int | None, summary: str) -> bool:
    lodging_payload = object_value(payload, "lodging")
    if lodging_payload.get("requested") is False:
        return True
    visit_type = string_value(first_value(payload, "visitType", "visit_type", "requestType", "request_type", "type"))
    if visit_type and visit_type.lower().replace("-", "_") in {"birding_day_trip", "day_trip", "day_visit", "pasadia", "pasadía"}:
        return True
    if arrival and departure and arrival == departure:
        return True
    if nights == 0:
        return True
    no_lodging = bool(re.search(r"\b(no cabin|no lodging|no hospedaje|sin caba[ñn]a|sin hospedaje)\b", summary, re.I))
    return bool(re.search(r"\b(day trip|day visit|bird(?:ing)? tour|tour de aves|pajareo|pasad[ií]a)\b", summary, re.I) and (no_lodging or not re.search(r"\b(cabin|caba[ñn]a|lodging|hospedaje|night|noche)\b", summary, re.I)))


def infer_birding_days(payload: dict[str, Any], nights: int | None, summary: str) -> int:
    birding_payload = object_value(payload, "birding")
    value = coerce_int(
        first_value(payload, "birdingDays", "birding_days", "birdTourDays", "bird_tour_days", "morningTourDays", "morning_tour_days"),
        None,
    )
    if value is None:
        value = coerce_int(birding_payload.get("morningTourDays"), None)
    if value is None:
        value = coerce_int(birding_payload.get("morning_tours"), None)
    if value is None:
        value = coerce_int(birding_payload.get("morningTours"), None)
    if value is None:
        for key in ("days", "dayCount", "day_count", "tourDays", "tour_days", "birdingDays", "birding_days"):
            value = coerce_int(birding_payload.get(key), None)
            if value is not None:
                break
    if value is None:
        for item in item_codes(payload):
            code = item_code_text(item)
            if "bird" in code or "ave" in code or "pajareo" in code:
                days = item.get("days")
                if isinstance(days, str) and days.lower() in {"each_day", "daily", "every_day"}:
                    value = nights or 1
                else:
                    value = coerce_int(days, None)
                if value is None:
                    value = coerce_int(item.get("count"), None)
                if value is None:
                    value = coerce_int(item.get("nights"), None)
                if value is None:
                    value = coerce_int(item.get("quantity"), None)
                break
    if value is None and re.search(r"\b(birding every day|daily birding|bird tour each (?:morning|day)|pajareo todos)\b", summary, re.I):
        value = nights or 1
    if value is None:
        visit_type = string_value(first_value(payload, "visitType", "visit_type", "type"))
        if visit_type and visit_type.lower().replace("-", "_") in {"birding_day_trip", "birding_tour_only"}:
            value = 1
        elif re.search(r"\b(bird(?:ing)? tour|tour de aves|avistamiento de aves|pajareo)\b", summary, re.I):
            value = 1
    return max(0, value or 0)


def infer_transport_requested(payload: dict[str, Any], summary: str) -> bool:
    transport_payload = object_value(payload, "transport")
    if isinstance(transport_payload.get("requested"), bool):
        return bool(transport_payload.get("requested"))
    if first_value(payload, "transport_included", "transportIncluded") is False:
        return False
    value = string_value(first_value(payload, "transport", "transportRequest", "transport_request"))
    if value and value.lower() in {"none", "no", "excluded", "sin transporte"}:
        return False
    lowered_summary = summary.lower()
    if "no transport" in lowered_summary or "sin transporte" in lowered_summary:
        return False
    return bool(value or "transport" in lowered_summary or "transporte" in lowered_summary)


def infer_bilingual_guide(payload: dict[str, Any], summary: str) -> str:
    birding_payload = object_value(payload, "birding")
    raw = birding_payload.get("bilingualGuide") or birding_payload.get("bilingual_guide") or first_value(payload, "bilingualGuide", "bilingual_guide")
    if isinstance(raw, bool):
        return "half_day" if raw else "none"
    if isinstance(raw, str):
        lowered = raw.lower()
        if lowered in {"half_day", "half day", "half-day", "medio dia", "medio día"}:
            return "half_day"
        if lowered in {"full_day", "full day", "full-day", "dia completo", "día completo"}:
            return "full_day"
        if lowered in {"none", "no", "local", "local_spanish", "spanish", "español", "espanol"}:
            return "none"
    if re.search(r"\b(bilingual|biling[uü]e)\b", summary, re.I):
        return "half_day"
    return "none"


def normalize_calculate_payload(payload: dict[str, Any]) -> dict[str, Any]:
    summary = request_summary(payload)
    arrival = date_value(first_value(payload, "arrivalDate", "arrival_date", "arrival", "startDate", "start_date", "date", "visitDate", "visit_date", "checkIn", "check_in"))
    departure = date_value(first_value(payload, "departureDate", "departure_date", "departure", "endDate", "end_date", "date", "visitDate", "visit_date", "checkOut", "check_out"))
    if not arrival:
        match = re.search(r"\b(?:arrival|arrive|check[ -]?in|llegada)\D*(\d{4}[-/]\d{2}[-/]\d{2})", summary, re.I)
        arrival = date_value(match.group(1)) if match else None
    if not departure:
        match = re.search(r"\b(?:departure|depart|check[ -]?out|salida)\D*(\d{4}[-/]\d{2}[-/]\d{2})", summary, re.I)
        departure = date_value(match.group(1)) if match else None
    nights = infer_nights(payload, arrival, departure, summary)
    guest_count = infer_guest_count(payload, summary)
    lodging_payload = object_value(payload, "lodging")
    meals_payload = object_value(payload, "meals")
    birding_payload = object_value(payload, "birding")
    transport_payload = object_value(payload, "transport")
    full_board = has_full_board(payload, summary)
    cabin_count = infer_cabin_count(payload, summary)
    birding_days = infer_birding_days(payload, nights, summary)
    day_trip = is_day_trip_quote(payload, arrival, departure, nights, summary)
    staff_count = infer_staff_count(payload, "guide") + infer_staff_count(payload, "driver")

    normalized: dict[str, Any] = {
        "propertyId": string_value(first_value(payload, "propertyId", "property_id")) or "owlswatch",
        "year": rate_year_from_payload({**payload, "arrivalDate": arrival, "departureDate": departure}),
    }
    audience = infer_audience(payload, summary)
    if audience:
        normalized["audience"] = audience
    if arrival:
        normalized["arrivalDate"] = arrival
    if departure:
        normalized["departureDate"] = departure
    if nights is not None:
        normalized["nights"] = nights

    children = coerce_int(object_value(payload, "guests").get("children"), 0) or 0
    children_under_two = coerce_int(object_value(payload, "guests").get("childrenUnderTwo"), None)
    if children_under_two is None:
        children_under_two = coerce_int(object_value(payload, "guests").get("children_under_two"), 0) or 0
    guides = coerce_int(first_value(payload, "guideCount", "guide_count"), None)
    if guides is None:
        guides = coerce_int(object_value(payload, "guests").get("guides"), 0) or 0
    if guides is None or guides == 0:
        birding_guides = coerce_int(object_value(payload, "birding").get("guideCount"), None)
        if birding_guides is None:
            birding_guides = coerce_int(object_value(payload, "birding").get("guide_count"), None)
        if birding_guides is not None:
            guides = birding_guides
    if guest_count is not None:
        normalized["guests"] = {
            "adults": max(0, guest_count - children),
            "children": children,
            "childrenUnderTwo": children_under_two,
            "guides": guides,
        }

    explicit_lodging_requested = lodging_payload.get("requested")
    if isinstance(explicit_lodging_requested, bool):
        lodging_requested = explicit_lodging_requested
    else:
        lodging_cabins = coerce_int(lodging_payload.get("cabins"), None)
        lodging_nights = coerce_int(lodging_payload.get("nights"), None)
        lodging_requested = not (day_trip or lodging_cabins == 0 or lodging_nights == 0)
    guide_room_count = coerce_int(lodging_payload.get("guideRoomCount"), None)
    if guide_room_count is None:
        guide_room_count = coerce_int(lodging_payload.get("guide_room_count"), None)
    if guide_room_count is None:
        guide_room_count = coerce_int(first_value(payload, "guideRoomCount", "guide_room_count"), None)
    if guide_room_count is None and lodging_requested and requests_guide_room(payload, summary):
        guide_room_count = 1
    normalized["lodging"] = {
        "requested": lodging_requested,
        "cabinCount": 0 if not lodging_requested else cabin_count,
        "cabinPreference": string_value(lodging_payload.get("cabinPreference") or lodging_payload.get("cabin_preference")) or "forest",
    }
    if guide_room_count:
        normalized["lodging"]["guideRoomCount"] = max(0, guide_room_count)

    lunches = nested_meal_count(meals_payload, "lunch")
    if lunches is None:
        lunches = item_meal_count(payload, "lunch")
    dinners = nested_meal_count(meals_payload, "dinner")
    if dinners is None:
        dinners = item_meal_count(payload, "dinner")
    if lunches is None:
        lunches = coerce_int(first_value(payload, "lunchCount", "lunch_count"), None)
    if dinners is None:
        dinners = coerce_int(first_value(payload, "dinnerCount", "dinner_count"), None)
    if full_board:
        lunches = lunches if lunches is not None else default_cabin_lunch_count(nights, summary)
        dinners = dinners if dinners is not None else (nights or 0)
    elif day_trip:
        lunches = normalize_day_trip_meal_count(lunches, guest_count, staff_count)
        dinners = normalize_day_trip_meal_count(dinners, guest_count, staff_count)
        if lunches is None and (requests_lunch(payload, summary) or birding_days > 0):
            lunches = 1
        dinners = dinners if dinners is not None else 0
    else:
        if lunches is None and requests_lunch(payload, summary):
            lunches = default_cabin_lunch_count(nights, summary) if normalized["lodging"]["requested"] else (nights or 1)
        if dinners is None and requests_dinner(payload, summary):
            dinners = nights or 1
    breakfasts = nested_meal_count(meals_payload, "breakfast")
    if breakfasts is None:
        breakfasts = item_meal_count(payload, "breakfast")
    if day_trip:
        breakfasts = normalize_day_trip_meal_count(breakfasts, guest_count, staff_count)
        if breakfasts is None:
            breakfasts = 1 if requests_breakfast(payload, summary) else 0
    else:
        breakfasts = breakfasts or "included_with_lodging"
    normalized["meals"] = {
        "breakfasts": breakfasts,
        "lunches": max(0, lunches or 0),
        "dinners": max(0, dinners or 0),
    }

    normalized["birding"] = {
        "morningTourDays": birding_days,
        "afternoonExtensionDays": coerce_int(birding_payload.get("afternoonExtensionDays"), None)
        or coerce_int(birding_payload.get("afternoon_extension_days"), 0)
        or 0,
        "bilingualGuide": infer_bilingual_guide(payload, summary),
    }

    normalized["transport"] = {
        "requested": infer_transport_requested(payload, summary),
    }
    if "notes" in transport_payload:
        normalized["transport"]["notes"] = transport_payload["notes"]

    return normalized


def upper_enum(value: Any) -> str | None:
    text = string_value(value)
    if not text:
        return None
    return re.sub(r"[^A-Z0-9_]+", "_", text.strip().upper()).strip("_") or None


def source_type_value(value: Any) -> str | None:
    normalized = upper_enum(value)
    if not normalized:
        return None
    mapping = {
        "TELEGRAM_GROUP": "TELEGRAM",
        "TELEGRAM_TOPIC": "TELEGRAM",
        "GMAIL_THREAD": "GMAIL",
        "EMAIL": "GMAIL",
        "WHATSAPP": "WHATSAPP_PASTE",
        "WHATSAPP_TEXT": "WHATSAPP_PASTE",
    }
    return mapping.get(normalized, normalized)


def pick_nested_object(payload: dict[str, Any], *keys: str) -> dict[str, Any]:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, dict):
            return value
    return {}


def nested_first(payload: dict[str, Any], nested: dict[str, Any], *keys: str) -> Any:
    value = first_value(payload, *keys)
    if value is not None:
        return value
    for key in keys:
        if key in nested:
            return nested[key]
    return None


def list_value(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def normalize_id_part(value: Any) -> str:
    text = string_value(value) or "unknown"
    text = re.sub(r"^telegram:", "", text)
    return re.sub(r"[^A-Za-z0-9-]+", "-", text).strip("-") or "unknown"


def source_id_from_trace(source_type: str, trace: dict[str, Any]) -> str:
    if source_type == "TELEGRAM":
        chat = trace.get("chat_id") or trace.get("chatId") or trace.get("sourceChatId")
        message = trace.get("message_id") or trace.get("messageId") or trace.get("sourceMessageId")
        if chat and message:
            return f"telegram-{normalize_id_part(chat)}-{normalize_id_part(message)}"
    if source_type == "GMAIL":
        thread = trace.get("threadId") or trace.get("thread_id") or trace.get("gmailThreadId")
        if thread:
            return f"gmail-thread-{normalize_id_part(thread)}"
    return f"{source_type.lower()}-{hashlib.sha256(json_dumps(trace).encode()).hexdigest()[:16]}"


def default_source_trace(payload: dict[str, Any]) -> dict[str, Any]:
    trace = pick_nested_object(payload, "sourceTrace", "source_trace")
    if trace:
        return dict(trace)
    source = payload.get("source")
    if isinstance(source, dict):
        return dict(source)
    result: dict[str, Any] = {}
    for key in ("sourceChatId", "sourceMessageId", "sourceTopicId", "sourceThreadId", "chatId", "messageId", "topicId"):
        if key in payload:
            result[key] = payload[key]
    return result


def calculation_payload(payload: dict[str, Any]) -> dict[str, Any]:
    calculation = validate_object("calculation", payload.get("calculation")) if isinstance(payload.get("calculation"), dict) else {}
    if calculation:
        return calculation
    line_items = list_value(payload.get("lineItems"))
    total = coerce_int(payload.get("totalCop"), None)
    if line_items or total is not None:
        return {
            "currency": string_value(payload.get("currency")) or "COP",
            "pricebookVersion": string_value(payload.get("pricebookVersion")),
            "lineItems": line_items,
            "subtotalCop": coerce_int(payload.get("subtotalCop"), total or 0) or 0,
            "discountCop": coerce_int(payload.get("discountCop"), 0) or 0,
            "totalCop": total or 0,
        }
    raise ToolError("invalid_input", "calculation is required to create a quote draft.")


def requested_services_payload(payload: dict[str, Any], request: dict[str, Any], normalized_calc: dict[str, Any]) -> dict[str, Any]:
    requested = payload.get("requestedServices")
    if isinstance(requested, dict):
        return requested
    if isinstance(request, dict) and request:
        return request
    services: dict[str, Any] = {
        "lodging": normalized_calc.get("lodging", {}),
        "meals": normalized_calc.get("meals", {}),
        "birding": normalized_calc.get("birding", {}),
        "transport": normalized_calc.get("transport", {}),
    }
    return {key: value for key, value in services.items() if value}


def normalize_quote_create_payload(payload: dict[str, Any]) -> dict[str, Any]:
    request = pick_nested_object(payload, "request", "parsedRequest", "parsed_request")
    source_trace = default_source_trace(payload)
    summary = (
        string_value(nested_first(payload, request, "requestSummary", "summary", "rawRequestSummary"))
        or request_summary(payload)
        or "Quote request"
    )
    normalized_calc = normalize_calculate_payload({**request, **payload, "requestSummary": summary})
    calculation = apply_quote_display_policies({**request, **payload}, calculation_payload(payload))

    audience = infer_audience({**request, **payload}, summary)
    if not audience:
        raise ToolError("missing_field", "Audience is required before creating a quote draft.")

    status = upper_enum(payload.get("status")) or ("NEEDS_INFO" if list_value(payload.get("missingFields")) else "DRAFTED")
    if status in {"SENT", "ACCEPTED", "APPROVED", "READY_TO_SEND", "REJECTED", "CANCELLED"}:
        raise ToolError("forbidden_status", "Cotiza can only create draft quote rows.")
    if status not in {"DRAFTED", "NEEDS_INFO"}:
        raise ToolError("invalid_input", "Quote status must be DRAFTED or NEEDS_INFO.")

    source_type = (
        source_type_value(payload.get("sourceType"))
        or source_type_value(payload.get("source"))
        or source_type_value(source_trace.get("channel"))
        or "TELEGRAM"
    )
    source_id = (
        string_value(payload.get("sourceId"))
        or string_value(payload.get("sourceMessageId"))
        or source_id_from_trace(source_type, source_trace)
    )
    idempotency_key = (
        string_value(payload.get("idempotencyKey"))
        or string_value(payload.get("idempotency_key"))
        or source_id
    )

    missing_fields = list_value(payload.get("missingFields")) or list_value(calculation.get("missingFields"))
    assumptions = list_value(payload.get("assumptions")) or list_value(calculation.get("assumptions"))
    guest_count = infer_guest_count({**request, **payload}, summary)
    guide_count = coerce_int(nested_first(payload, request, "guideCount", "guide_count"), 0) or 0

    return {
        "propertyId": string_value(payload.get("propertyId")) or "owlswatch",
        "sourceType": source_type,
        "sourceId": source_id,
        "sourceUrl": string_value(payload.get("sourceUrl") or source_trace.get("sourceUrl")),
        "sourceReceivedAt": string_value(payload.get("sourceReceivedAt")) or now_iso(),
        "requesterName": string_value(nested_first(payload, request, "requesterName", "requester_name")),
        "requesterEmail": string_value(nested_first(payload, request, "requesterEmail", "requester_email")),
        "requesterPhone": string_value(nested_first(payload, request, "requesterPhone", "requester_phone")),
        "agencyName": string_value(nested_first(payload, request, "agencyName", "agency_name", "operatorName", "operator_name")),
        "clientName": string_value(nested_first(payload, request, "clientName", "client_name")),
        "audience": "OPERATOR" if audience == "operator" else "DIRECT",
        "language": string_value(payload.get("language")) or "en",
        "arrivalDate": date_value(nested_first(payload, request, "arrivalDate", "arrival_date", "arrival", "startDate", "start_date", "date", "visitDate", "visit_date")),
        "departureDate": date_value(nested_first(payload, request, "departureDate", "departure_date", "departure", "endDate", "end_date", "date", "visitDate", "visit_date")),
        "nights": infer_nights({**request, **payload}, normalized_calc.get("arrivalDate"), normalized_calc.get("departureDate"), summary),
        "guestCount": guest_count,
        "guideCount": guide_count,
        "requestSummary": summary,
        "rawRequestText": string_value(payload.get("rawRequestText") or payload.get("rawRequest") or payload.get("rawText")) or summary,
        "requestedServices": requested_services_payload(payload, request, normalized_calc),
        "calculation": calculation,
        "assumptions": assumptions,
        "missingFields": missing_fields,
        "sourceTrace": source_trace,
        "replyDraft": string_value(payload.get("replyDraft")) or "Draft only; do not send.",
        "agentMetadata": payload.get("agentMetadata") if isinstance(payload.get("agentMetadata"), dict) else {"agent": "Cotiza", "workflow": "quote-draft"},
        "idempotencyKey": idempotency_key,
        "status": status,
    }


MONTH_NAME_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?|"
    r"enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre"
    r")\.?\s+(\d{1,2})(?:\s*[-–]\s*(\d{1,2}))?(?:\s*(?:,|/|de)?\s*)(\d{4})\b",
    re.I,
)

MONTH_LOOKUP = {
    "jan": 1, "january": 1, "enero": 1,
    "feb": 2, "february": 2, "febrero": 2,
    "mar": 3, "march": 3, "marzo": 3,
    "apr": 4, "april": 4, "abril": 4,
    "may": 5, "mayo": 5,
    "jun": 6, "june": 6, "junio": 6,
    "jul": 7, "july": 7, "julio": 7,
    "aug": 8, "august": 8, "agosto": 8,
    "sep": 9, "sept": 9, "september": 9, "septiembre": 9, "setiembre": 9,
    "oct": 10, "october": 10, "octubre": 10,
    "nov": 11, "november": 11, "noviembre": 11,
    "dec": 12, "december": 12, "diciembre": 12,
}


def month_number(token: str | None) -> int | None:
    if not token:
        return None
    normalized = token.lower().rstrip(".")
    return MONTH_LOOKUP.get(normalized) or MONTH_LOOKUP.get(normalized[:3])


def compact_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def merge_dicts(*objects: dict[str, Any]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for obj in objects:
        if not isinstance(obj, dict):
            continue
        for key, value in obj.items():
            if isinstance(value, dict) and isinstance(merged.get(key), dict):
                merged[key] = merge_dicts(merged[key], value)
            elif value is not None:
                merged[key] = value
    return merged


def parse_text_dates(text: str) -> tuple[str | None, str | None]:
    iso_dates = re.findall(r"\b\d{4}-\d{2}-\d{2}\b", text)
    if len(iso_dates) >= 2:
        return iso_dates[0], iso_dates[1]
    if len(iso_dates) == 1:
        return iso_dates[0], iso_dates[0]

    month_words = (
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?|"
        r"enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre"
    )

    cross_month = re.search(
        rf"\b({month_words})\.?\s+(\d{{1,2}})(?:st|nd|rd|th)?\s*[-–]\s*({month_words})\.?\s+(\d{{1,2}})(?:st|nd|rd|th)?(?:\s*,)?\s+(\d{{4}})\b",
        text,
        re.I,
    )
    if cross_month:
        start_month, start_day, end_month, end_day, year = cross_month.groups()
        sm = month_number(start_month)
        em = month_number(end_month)
        if sm and em:
            return dt.date(int(year), sm, int(start_day)).isoformat(), dt.date(int(year), em, int(end_day)).isoformat()

    spanish_day_range = re.search(
        rf"\b(\d{{1,2}})\s*[-–]\s*(\d{{1,2}})\s+de\s+({month_words})\.?(?:\s*,)?\s+(\d{{4}})\b",
        text,
        re.I,
    )
    if spanish_day_range:
        start_day, end_day, month_token, year = spanish_day_range.groups()
        month = month_number(month_token)
        if month:
            return dt.date(int(year), month, int(start_day)).isoformat(), dt.date(int(year), month, int(end_day)).isoformat()

    day_month_dates: list[str] = []
    for day, month_token, year in re.findall(
        rf"\b(\d{{1,2}})\s+(?:de\s+)?({month_words})\.?(?:\s+de)?\s+(\d{{4}})\b",
        text,
        re.I,
    ):
        month = month_number(month_token)
        if month:
            day_month_dates.append(dt.date(int(year), month, int(day)).isoformat())
    if len(day_month_dates) >= 2:
        return day_month_dates[0], day_month_dates[1]
    if len(day_month_dates) == 1:
        return day_month_dates[0], day_month_dates[0]

    match = MONTH_NAME_RE.search(text)
    if match:
        month_token, start_day, end_day, year = match.groups()
        month = month_number(month_token)
        if month:
            start = dt.date(int(year), month, int(start_day))
            end = dt.date(int(year), month, int(end_day or start_day))
            return start.isoformat(), end.isoformat()

    numeric = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b", text)
    if numeric:
        first, second, year = numeric.groups()
        try:
            date = dt.date(int(year), int(first), int(second))
        except ValueError:
            date = dt.date(int(year), int(second), int(first))
        return date.isoformat(), date.isoformat()
    return None, None


def extract_named_value(text: str, labels: tuple[str, ...]) -> str | None:
    pattern = "|".join(re.escape(label) for label in labels)
    match = re.search(rf"\b(?:{pattern})\b\s*[:\-]?\s*([^\n,.;]+)", text, re.I)
    if not match:
        return None
    value = compact_text(match.group(1))
    value = re.split(r"\b(?:for client|client|guest|fecha|date|guests?|clients?|pax)\b", value, flags=re.I)[0].strip()
    return value or None


def raw_text_quote_intent(raw_text: str) -> dict[str, Any]:
    lowered = raw_text.lower()
    arrival, departure = parse_text_dates(raw_text)
    intent: dict[str, Any] = {
        "requestSummary": compact_text(raw_text),
        "rawRequestText": raw_text,
    }
    if arrival:
        intent["arrivalDate"] = arrival
    if departure:
        intent["departureDate"] = departure

    audience = infer_audience({}, raw_text)
    if not audience and re.search(r"\bestado\s+reservaci[oó]n\b|\breferencia\b.{0,80}\bservicio\b|\bhotel\s+owl", raw_text, re.I):
        audience = "operator"
    if audience:
        intent["audience"] = audience

    agency = extract_named_value(raw_text, ("operator", "operador", "agency", "agencia"))
    if agency:
        intent["agencyName"] = agency
    client = extract_named_value(raw_text, ("client", "cliente", "guest", "huesped", "huésped", "referencia", "reference"))
    client_match = re.search(r"\bfor client\s+([^\n,.;]+)", raw_text, re.I)
    if client_match:
        client = compact_text(client_match.group(1))
    if client:
        intent["clientName"] = client

    guest_match = re.search(r"\b(\d{1,3})\s*(?:clients?|guests?|pax|people|persons?|personas?|adults?|adultos?)\b", raw_text, re.I)
    if guest_match:
        intent["guestCount"] = int(guest_match.group(1))
    guide_match = re.search(r"\b(\d{1,2})\s*(?:guides?|gu[ií]as?)\b", raw_text, re.I)
    if guide_match:
        intent["guideCount"] = int(guide_match.group(1))
    elif re.search(r"(?:\+|plus|con)\s*(?:1\s*)?gu[ií]a\b|\bgu[ií]a\b", raw_text, re.I):
        intent["guideCount"] = 1
    driver_match = re.search(r"\b(\d{1,2})\s*(?:drivers?|conductores?|chofer(?:es)?)\b", raw_text, re.I)
    if driver_match:
        intent["driverCount"] = int(driver_match.group(1))

    day_trip = bool(re.search(r"\b(birding day trip|bird(?:ing)? tour|day trip|day visit|tour de aves|avistamiento de aves|pajareo|pasad[ií]a)\b", raw_text, re.I))
    no_lodging = bool(re.search(r"\b(no cabin|no lodging|no hospedaje|sin caba[ñn]a|sin hospedaje)\b", raw_text, re.I))
    cabin = bool(re.search(r"\b(cabin|caba[ñn]a|lodging|hospedaje|hotel|habitaci[oó]n|matrimonial|single|nights?|noches?)\b", raw_text, re.I)) and not no_lodging
    if day_trip and not cabin:
        intent["visitType"] = "birding_day_trip"
        intent["lodging"] = {"requested": False, "cabinCount": 0}
        intent["birding"] = {"morningTourDays": 1}
    elif no_lodging:
        intent["lodging"] = {"requested": False, "cabinCount": 0}
    elif cabin:
        intent["visitType"] = "cabin"
        intent["lodging"] = {"requested": True, "cabinCount": infer_cabin_count({"requestSummary": raw_text}, raw_text)}
        if requests_guide_room(intent, raw_text):
            intent["lodging"]["guideRoomCount"] = 1

    meals: dict[str, Any] = {}
    if re.search(r"\b(full board|food included|meals included|comida incluida|pensi[oó]n completa|a?alimentaci[oó]n completa)\b", raw_text, re.I):
        meals["plan"] = "full_board"
    elif re.search(r"\b(breakfast|desayuno)\b", raw_text, re.I) and re.search(r"\b(lunch|almuerzo)\b", raw_text, re.I):
        meals["plan"] = "breakfast_and_lunch"
    else:
        if re.search(r"\b(breakfast|desayuno)\b", raw_text, re.I):
            meals["breakfast"] = True
        if re.search(r"\b(lunch|almuerzo)\b", raw_text, re.I):
            meals["lunch"] = True
        if re.search(r"\b(dinners?|cenas?)\b", raw_text, re.I):
            meals["dinner"] = True
    if meals:
        intent["meals"] = meals

    if "no transport" in lowered or "sin transporte" in lowered:
        intent["transport"] = {"requested": False}
    elif "transport" in lowered or "transporte" in lowered:
        intent["transport"] = {"requested": True}

    return intent


def prepare_quote_intent(args: dict[str, Any]) -> dict[str, Any]:
    raw_text = validate_text("raw_text", args.get("raw_text"), required=True, max_len=20000) or ""
    parsed_intent = validate_object("parsed_intent", args.get("parsed_intent")) if isinstance(args.get("parsed_intent"), dict) else {}
    prior_context = validate_object("prior_context", args.get("prior_context")) if isinstance(args.get("prior_context"), dict) else {}
    user_overrides = validate_object("user_overrides", args.get("user_overrides")) if isinstance(args.get("user_overrides"), dict) else {}
    return merge_dicts(prior_context, raw_text_quote_intent(raw_text), parsed_intent, user_overrides, {"rawRequestText": raw_text})


def quote_prepare_missing(intent: dict[str, Any], normalized: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    summary = request_summary(intent)
    audience = infer_audience(intent, summary)
    if not audience:
        missing.append("operator/direct")
    if not date_value(normalized.get("arrivalDate")):
        missing.append("dates/year")
    if infer_guest_count(intent, summary) is None:
        missing.append("guest count")
    visit_type = string_value(first_value(intent, "visitType", "visit_type", "requestType", "request_type"))
    lodging_requested = object_value(intent, "lodging").get("requested")
    birding_days = coerce_int(object_value(normalized, "birding").get("morningTourDays"), 0) or 0
    if not (visit_type or isinstance(lodging_requested, bool) or birding_days > 0 or sheet_has_visit_type({**intent, "calculation": {"lineItems": []}})):
        missing.append("cabin stay or birding day trip")
    if string_value(first_value(intent, "visitType", "visit_type")) == "cabin" and infer_nights(intent, normalized.get("arrivalDate"), normalized.get("departureDate"), summary) is None:
        missing.append("number of nights")
    return list(dict.fromkeys(missing))


def question_for_missing(missing: list[str]) -> str:
    if missing == ["operator/direct"]:
        return "Is this for an operator/agency quote or a direct client quote?"
    if missing == ["operator"]:
        return "Who is the operator or agency?"
    if missing == ["dates/year"]:
        return "What are the quote dates, including year?"
    if missing == ["guest count"]:
        return "How many guests or clients?"
    if missing == ["cabin stay or birding day trip"]:
        return "Is this a cabin stay or a birding day trip?"
    return "I can draft this, but I need: " + ", ".join(missing) + "."


def known_quote_details(intent: dict[str, Any], normalized: dict[str, Any]) -> dict[str, Any]:
    return {
        "audience": normalized.get("audience"),
        "agencyName": string_value(first_value(intent, "agencyName", "agency_name", "operatorName", "operator_name")),
        "clientName": string_value(first_value(intent, "clientName", "client_name")),
        "arrivalDate": normalized.get("arrivalDate"),
        "departureDate": normalized.get("departureDate"),
        "guestCount": infer_guest_count(intent, request_summary(intent)),
        "visitType": string_value(first_value(intent, "visitType", "visit_type", "requestType", "request_type")),
        "meals": normalized.get("meals"),
        "birding": normalized.get("birding"),
        "lodging": normalized.get("lodging"),
        "transport": normalized.get("transport"),
    }


def build_prepared_quote(args: dict[str, Any]) -> dict[str, Any]:
    raw_text = validate_text("raw_text", args.get("raw_text"), required=True, max_len=20000) or ""
    source_metadata = validate_object("source_metadata", args.get("source_metadata")) if isinstance(args.get("source_metadata"), dict) else {}
    intent = prepare_quote_intent(args)
    normalized = normalize_calculate_payload(intent)
    missing = quote_prepare_missing(intent, normalized)
    if missing:
        return {
            "ok": True,
            "status": "needs_info",
            "question": question_for_missing(missing),
            "missing": missing,
            "known": known_quote_details(intent, normalized),
            "safeDefaults": [
                "Client/guest name may be left blank.",
                "Breakfast requested for a day trip is priced for clients; guide/driver breakfast is complimentary.",
                "Transport is excluded unless requested.",
                "Bilingual guide is excluded unless requested.",
            ],
        }

    calculation = tool_quote_calculate({"payload": intent})
    if not calculation.get("ok"):
        return {"ok": True, "status": "calculation_failed", "error": calculation.get("error")}
    if infer_audience(intent, request_summary(intent)) == "operator" and not string_value(first_value(intent, "agencyName", "agency_name", "operatorName", "operator_name")):
        missing_fields = list_value(calculation.get("missingFields"))
        if "operator name" not in [str(item).lower() for item in missing_fields]:
            calculation = {**calculation, "missingFields": [*missing_fields, "operator name"]}
    line_items = calculation.get("lineItems")
    if not isinstance(line_items, list) or not line_items:
        return {
            "ok": True,
            "status": "needs_info",
            "question": "I could not identify the billable quote items. Is this a cabin stay or a birding day trip?",
            "missing": ["billable items"],
            "known": known_quote_details(intent, normalized),
            "calculation": calculation,
        }

    prepared_quote = {
        "version": 1,
        "rawText": raw_text,
        "sourceMetadata": source_metadata,
        "intent": intent,
        "canonicalPayload": normalized,
        "calculation": calculation,
        "assumptions": list_value(calculation.get("assumptions")),
        "missingOptional": list_value(calculation.get("missingFields")),
        "warnings": list_value(calculation.get("flags")),
        "requestSummary": request_summary(intent),
    }
    return {
        "ok": True,
        "status": "ready_preview",
        "preparedQuote": prepared_quote,
        "canonicalPayload": normalized,
        "calculation": calculation,
        "assumptions": prepared_quote["assumptions"],
        "missingOptional": prepared_quote["missingOptional"],
        "warnings": prepared_quote["warnings"],
    }


def tool_quote_prepare(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"raw_text", "source_metadata", "prior_context", "user_overrides", "parsed_intent", "mode"})
    validate_text("mode", args.get("mode"), max_len=40)
    return build_prepared_quote(args)


def redo_requested(prepared: dict[str, Any], source_metadata: dict[str, Any], explicit: Any = None) -> bool:
    if bool_value(explicit):
        return True
    metadata = merge_dicts(prepared.get("sourceMetadata") if isinstance(prepared.get("sourceMetadata"), dict) else {}, source_metadata)
    for key in ("redo", "forceRedo", "force_redo", "recreate", "regenerate"):
        if bool_value(metadata.get(key)):
            return True
    text = " ".join([
        string_value(prepared.get("rawText")) or "",
        string_value(prepared.get("requestSummary")) or "",
        string_value(metadata.get("commandText")) or "",
        string_value(metadata.get("userText")) or "",
    ]).lower()
    return bool(re.search(r"\b(re-?do|recalculate|regenerate|recreate|fresh draft|new test draft|new draft from same source)\b", text))


def redo_idempotency_suffix(prepared: dict[str, Any], source_metadata: dict[str, Any]) -> str:
    metadata = merge_dicts(prepared.get("sourceMetadata") if isinstance(prepared.get("sourceMetadata"), dict) else {}, source_metadata)
    seed = (
        string_value(metadata.get("redoKey"))
        or string_value(metadata.get("redo_key"))
        or string_value(metadata.get("idempotencyNonce"))
        or string_value(metadata.get("idempotency_nonce"))
        or string_value(metadata.get("sourceMessageId"))
        or string_value(metadata.get("messageId"))
        or string_value(metadata.get("message_id"))
        or now_iso()
    )
    return normalize_id_part(seed)[:80]


def versioned_idempotency_key(base: str) -> str:
    return f"{base}-rules-{normalize_id_part(QUOTE_RULE_VERSION)}"


def prepared_idempotency_key(prepared: dict[str, Any], source_metadata: dict[str, Any], explicit: str | None, redo: bool = False) -> str:
    if explicit:
        return explicit
    metadata = merge_dicts(prepared.get("sourceMetadata") if isinstance(prepared.get("sourceMetadata"), dict) else {}, source_metadata)
    chat = metadata.get("chat_id") or metadata.get("chatId")
    message = metadata.get("message_id") or metadata.get("messageId")
    if chat and message:
        base = f"telegram-{normalize_id_part(chat)}-{normalize_id_part(message)}"
        base = versioned_idempotency_key(base)
        return f"{base}-redo-{redo_idempotency_suffix(prepared, metadata)}" if redo else base
    thread = metadata.get("threadId") or metadata.get("thread_id")
    if thread:
        base = f"gmail-thread-{normalize_id_part(thread)}"
        base = versioned_idempotency_key(base)
        return f"{base}-redo-{redo_idempotency_suffix(prepared, metadata)}" if redo else base
    raw = string_value(prepared.get("rawText")) or json_dumps(prepared.get("intent") or prepared)
    base = "quote-prepared-" + hashlib.sha256(raw.encode()).hexdigest()[:20]
    base = versioned_idempotency_key(base)
    return f"{base}-redo-{redo_idempotency_suffix(prepared, metadata)}" if redo else base


def create_payload_from_prepared(prepared: dict[str, Any], source_metadata: dict[str, Any], idempotency_key: str | None, redo: bool = False) -> dict[str, Any]:
    intent = validate_object("preparedQuote.intent", prepared.get("intent"))
    canonical = validate_object("preparedQuote.canonicalPayload", prepared.get("canonicalPayload"))
    calculation = validate_object("preparedQuote.calculation", prepared.get("calculation"))
    metadata = merge_dicts(prepared.get("sourceMetadata") if isinstance(prepared.get("sourceMetadata"), dict) else {}, source_metadata)
    summary = string_value(prepared.get("requestSummary")) or request_summary(intent)
    arrival = date_value(canonical.get("arrivalDate"))
    departure = date_value(canonical.get("departureDate"))
    audience = infer_audience(intent, summary)
    payload = {
        **intent,
        "source": metadata.get("source") or metadata.get("channel") or "telegram",
        "sourceTrace": metadata,
        "quoteType": audience or canonical.get("audience"),
        "agencyName": string_value(first_value(intent, "agencyName", "agency_name", "operatorName", "operator_name")),
        "clientName": string_value(first_value(intent, "clientName", "client_name")),
        "arrivalDate": arrival,
        "departureDate": departure,
        "guestCount": infer_guest_count(intent, summary),
        "requestSummary": summary,
        "rawRequestText": string_value(prepared.get("rawText")) or string_value(intent.get("rawRequestText")) or summary,
        "requestedServices": {
            "lodging": canonical.get("lodging", {}),
            "meals": canonical.get("meals", {}),
            "birding": canonical.get("birding", {}),
            "transport": canonical.get("transport", {}),
        },
        "calculation": calculation,
        "assumptions": list_value(prepared.get("assumptions")) or list_value(calculation.get("assumptions")),
        "missingFields": list_value(prepared.get("missingOptional")) or list_value(calculation.get("missingFields")),
        "sourceTrace": metadata,
        "replyDraft": "Draft only; do not send.",
        "idempotencyKey": prepared_idempotency_key(prepared, metadata, idempotency_key, redo=redo),
        "status": "DRAFTED",
    }
    payload["agentMetadata"] = {
        "agent": "Cotiza",
        "workflow": "quote-draft",
        "quoteRuleVersion": QUOTE_RULE_VERSION,
    }
    if redo:
        payload["agentMetadata"].update({
            "redo": True,
            "redoReason": string_value(metadata.get("redoReason") or metadata.get("redo_reason")) or "User requested a fresh draft from the same source.",
        })
    return payload


def create_draft_row(payload: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    if mocks_enabled(config):
        quote_id, quote_number = mock_quote_id(payload)
        MOCK_QUOTES_DIR.mkdir(parents=True, exist_ok=True)
        path = MOCK_QUOTES_DIR / f"{quote_id}.json"
        existing = None
        if path.exists():
            existing = json.loads(path.read_text())
        record = existing or {
            "quoteId": quote_id,
            "quoteNumber": quote_number,
            "status": payload.get("status") or "DRAFTED",
            "createdAt": now_iso(),
            "payload": payload,
        }
        record["updatedAt"] = now_iso()
        path.write_text(json.dumps(record, indent=2, ensure_ascii=False) + "\n")
        return {
            "ok": True,
            "mock": True,
            "idempotent": existing is not None,
            "quoteId": quote_id,
            "quoteNumber": quote_number,
            "status": record["status"],
            "reviewUrl": f"{operations_base_url(config)}/quotes/{quote_id}",
        }
    url = f"{operations_base_url(config)}/api/quotes/intake"
    data = http_json("POST", url, payload, bearer_headers(config), timeout=45, attempts=2)
    if "ok" not in data:
        data["ok"] = True
    return data


def create_drive_for_draft(row: dict[str, Any], payload: dict[str, Any]) -> tuple[dict[str, Any] | None, list[str]]:
    warnings: list[str] = []
    existing_url = string_value(row.get("driveSheetUrl") or row.get("driveUrl") or row.get("drive_sheet_url"))
    existing_file_id = string_value(row.get("driveFileId") or row.get("drive_file_id"))
    if existing_url:
        return {"ok": True, "driveFileId": existing_file_id, "driveSheetUrl": existing_url, "existing": True}, warnings
    quote_id = string_value(row.get("quoteId") or row.get("quote_id") or row.get("id"))
    quote_number = string_value(row.get("quoteNumber") or row.get("quote_number") or row.get("number"))
    if not quote_id or not quote_number:
        return None, ["Operations draft did not return quote id/number for Drive sheet creation."]
    drive_args = {
        "quoteId": quote_id,
        "quoteNumber": quote_number,
        "agencyName": payload.get("agencyName"),
        "requesterName": payload.get("requesterName"),
        "clientName": payload.get("clientName"),
        "arrivalDate": payload.get("arrivalDate"),
        "departureDate": payload.get("departureDate"),
        "guestCount": payload.get("guestCount"),
        "requestSummary": payload.get("requestSummary"),
        "calculation": payload.get("calculation"),
        "assumptions": list_value(payload.get("assumptions")),
        "missingFields": list_value(payload.get("missingFields")),
        "replyDraft": string_value(payload.get("replyDraft")),
        "sourceTrace": payload.get("sourceTrace") if isinstance(payload.get("sourceTrace"), dict) else {},
    }
    try:
        drive_result = tool_drive_create_quote_sheet(drive_args)
        if drive_result.get("ok") and drive_result.get("driveFileId") and drive_result.get("driveSheetUrl"):
            try:
                tool_quote_update_drive({
                    "quoteId": quote_id,
                    "driveFileId": drive_result["driveFileId"],
                    "driveSheetUrl": drive_result["driveSheetUrl"],
                })
            except Exception:
                warnings.append("Drive sheet was created, but Operations could not be patched with the Drive link.")
        return drive_result, warnings
    except Exception:
        warnings.append("Operations quote row was created, but Drive sheet creation failed.")
        return None, warnings


def quote_ref_text(value: Any) -> str:
    text = validate_text("quote_ref", value, required=True, max_len=120) or ""
    match = re.search(r"\bQ-\d{4}-\d{4,}\b", text, re.I)
    return match.group(0).upper() if match else text.strip()


def quote_number_matches(candidate: Any, reference: str) -> bool:
    return isinstance(candidate, str) and candidate.strip().upper() == reference.strip().upper()


def fetch_quote_by_reference(reference: str, config: dict[str, Any]) -> dict[str, Any]:
    ref = quote_ref_text(reference)
    if mocks_enabled(config):
        if not MOCK_QUOTES_DIR.exists():
            raise ToolError("quote_not_found", "No mock quote drafts exist.")
        for path in MOCK_QUOTES_DIR.glob("*.json"):
            try:
                record = json.loads(path.read_text())
            except Exception:
                continue
            if quote_number_matches(record.get("quoteNumber"), ref) or quote_number_matches(record.get("quoteId"), ref):
                payload = record.get("payload") if isinstance(record.get("payload"), dict) else {}
                return {
                    **payload,
                    "id": record.get("quoteId"),
                    "quoteNumber": record.get("quoteNumber"),
                    "status": record.get("status"),
                    "calculation": payload.get("calculation") if isinstance(payload.get("calculation"), dict) else {},
                    "lineItems": (payload.get("calculation") or {}).get("lineItems") if isinstance(payload.get("calculation"), dict) else [],
                }
        raise ToolError("quote_not_found", "Quote draft was not found.")

    base = operations_base_url(config)
    headers = bearer_headers(config)
    if ref.startswith("cm") or ref.startswith("cl"):
        data = http_get_json(f"{base}/api/quotes/{urllib.parse.quote(ref)}", headers, timeout=45)
        quote = data.get("quote") if isinstance(data.get("quote"), dict) else data
        if isinstance(quote, dict) and quote.get("id"):
            return quote

    list_data = http_get_json(f"{base}/api/quotes", headers, timeout=45)
    quotes = list_data.get("quotes") if isinstance(list_data.get("quotes"), list) else []
    match = next((item for item in quotes if isinstance(item, dict) and (quote_number_matches(item.get("quoteNumber"), ref) or quote_number_matches(item.get("id"), ref))), None)
    if not match:
        raise ToolError("quote_not_found", "Quote draft was not found.")
    quote_id = string_value(match.get("id"))
    if not quote_id:
        raise ToolError("quote_not_found", "Quote draft was not found.")
    data = http_get_json(f"{base}/api/quotes/{urllib.parse.quote(quote_id)}", headers, timeout=45)
    quote = data.get("quote") if isinstance(data.get("quote"), dict) else data
    if not isinstance(quote, dict) or not quote.get("id"):
        raise ToolError("quote_not_found", "Quote draft was not found.")
    return quote


REVISION_ITEM_ALIASES: dict[str, tuple[str, ...]] = {
    "breakfast": ("breakfast", "desayuno"),
    "lunch": ("lunch", "almuerzo", "comida"),
    "dinner": ("dinner", "cena"),
    "bird_tour": ("bird tour", "birding", "aves", "pajareo", "photography and bird tour"),
    "cabin": ("cabin", "cabana", "cabaña", "lodging", "hospedaje"),
}


def revision_instruction_action(text: str) -> str | None:
    lowered = text.lower()
    if re.search(r"\b(remove|delete|subtract|take out|drop|exclude|without|no)\b", lowered):
        return "remove"
    if re.search(r"\b(add|include|plus)\b", lowered):
        return "add"
    return None


def revision_instruction_quantity(text: str, item_type: str, quote: dict[str, Any]) -> int | None:
    lowered = text.lower()
    aliases = REVISION_ITEM_ALIASES.get(item_type, ())
    for alias in aliases:
        escaped = re.escape(alias)
        before = re.search(rf"\b(\d{{1,3}})\s+(?:client\s+|guide\s+|driver\s+|staff\s+)?{escaped}s?\b", lowered)
        after = re.search(rf"\b{escaped}s?\s+(\d{{1,3}})\b", lowered)
        if before or after:
            return int((before or after).group(1))
    if item_type == "lunch" and re.search(r"\b(checkout|check-out|departure|salida)\b", lowered):
        return coerce_int(quote.get("guestCount"), None)
    if "all" in lowered or "todos" in lowered or "todas" in lowered:
        return -1
    cleaned = re.sub(r"\bQ-\d{4}-\d{4,}\b", " ", text, flags=re.I)
    number_match = re.search(r"\b(\d{1,3})\b", cleaned)
    if number_match:
        return int(number_match.group(1))
    return None


def revision_instruction_item_type(text: str) -> str | None:
    lowered = text.lower()
    for item_type, aliases in REVISION_ITEM_ALIASES.items():
        if any(alias in lowered for alias in aliases):
            return item_type
    return None


def revision_role(text: str) -> str:
    lowered = text.lower()
    if re.search(r"\b(guide|driver|staff|gu[ií]a|conductor|chofer)\b", lowered):
        return "staff"
    return "client"


def item_matches_revision(item: dict[str, Any], item_type: str, role: str) -> bool:
    text = line_text(item)
    is_staff = "guide_driver" in text or "guide/driver" in text or "staff_meal" in text
    if role == "staff" and not is_staff:
        return False
    if role == "client" and is_staff:
        return False
    return any(alias in text for alias in REVISION_ITEM_ALIASES[item_type])


def revision_line_description(item: dict[str, Any]) -> str:
    return display_description(item.get("description") or item.get("name") or item.get("serviceCode") or "line item")


def recalculate_revision_calculation(calc: dict[str, Any], original_discount: int | None = None) -> dict[str, Any]:
    items = [item for item in (calc.get("lineItems") or calc.get("items") or []) if isinstance(item, dict)]
    subtotal = sum(line_total(item) for item in items)
    discount = original_discount if original_discount is not None else money_value(calc.get("discountCop"))
    discount = max(0, min(discount, subtotal))
    updated = dict(calc)
    updated["lineItems"] = items
    updated["subtotalCop"] = subtotal
    updated["discountCop"] = discount
    updated["totalCop"] = subtotal - discount
    return updated


def apply_revision_instruction(quote: dict[str, Any], instruction: str) -> tuple[dict[str, Any], str]:
    action = revision_instruction_action(instruction)
    item_type = revision_instruction_item_type(instruction)
    if not action or not item_type:
        raise ToolError(
            "unsupported_revision_instruction",
            "I can revise simple line-item instructions such as: remove 2 lunches, add 2 dinners, or remove checkout lunch.",
        )
    quantity = revision_instruction_quantity(instruction, item_type, quote)
    if quantity is None:
        raise ToolError("revision_needs_quantity", "Please include the quantity to add or remove, for example: remove 2 lunches.")

    calc = dict(quote.get("calculation") if isinstance(quote.get("calculation"), dict) else {})
    raw_items = calc.get("lineItems") or quote.get("lineItems") or []
    items = [dict(item) for item in raw_items if isinstance(item, dict)]
    target_index = next((idx for idx, item in enumerate(items) if item_matches_revision(item, item_type, revision_role(instruction))), None)
    if target_index is None:
        raise ToolError("revision_line_not_found", "I could not find a matching line item to revise.")

    target = dict(items[target_index])
    current_qty = quantity_value(target.get("quantity"))
    unit_price = money_value(target.get("unitPriceCop"))
    if current_qty <= 0 or unit_price < 0:
        raise ToolError("revision_line_not_supported", "The matching line item does not have a revisable quantity.")

    if quantity == -1:
        new_qty = 0 if action == "remove" else current_qty
        delta = current_qty
    elif action == "remove":
        delta = min(float(quantity), current_qty)
        new_qty = current_qty - delta
    else:
        delta = float(quantity)
        new_qty = current_qty + delta

    if new_qty <= 0:
        items.pop(target_index)
    else:
        target["quantity"] = int(new_qty) if float(new_qty).is_integer() else new_qty
        target["totalCop"] = int(round(unit_price * new_qty))
        items[target_index] = target

    calc["lineItems"] = items
    updated = recalculate_revision_calculation(calc, original_discount=money_value(quote.get("discountCop") or calc.get("discountCop")))
    line_name = revision_line_description(target)
    verb = "Removed" if action == "remove" else "Added"
    amount = int(delta) if float(delta).is_integer() else delta
    return updated, f"{verb} {amount} {line_name} unit{'s' if amount != 1 else ''}."


def date_only(value: Any) -> str | None:
    text = string_value(value)
    if not text:
        return None
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return match.group(0) if match else date_value(text)


def revision_payload_from_quote(quote: dict[str, Any], instruction: str, source_metadata: dict[str, Any]) -> tuple[dict[str, Any], str]:
    status = upper_enum(quote.get("status")) or "DRAFTED"
    if status not in {"DRAFTED", "NEEDS_INFO", "NEW"}:
        raise ToolError("quote_not_editable", "Cotiza can only revise quotes that are still drafts.")

    calculation, revision_note = apply_revision_instruction(quote, instruction)
    quote_number = string_value(quote.get("quoteNumber")) or string_value(quote.get("quote_number")) or "quote"
    quote_id = string_value(quote.get("id") or quote.get("quoteId")) or quote_number
    revision_hash = hashlib.sha256(f"{quote_number}:{instruction}:{QUOTE_RULE_VERSION}".encode()).hexdigest()[:12]
    assumptions = list_value(quote.get("assumptions")) or list_value(calculation.get("assumptions"))
    assumptions = [*assumptions, f"Revision of {quote_number}: {revision_note}"]
    source_trace = quote.get("sourceTrace") if isinstance(quote.get("sourceTrace"), dict) else {}
    source_trace = {
        **source_trace,
        "revisionOfQuoteId": quote_id,
        "revisionOfQuoteNumber": quote_number,
        "revisionInstruction": instruction,
        "revisionSource": source_metadata,
    }
    payload = {
        "propertyId": "owlswatch",
        "sourceType": "MANUAL",
        "sourceId": f"revision-{normalize_id_part(quote_number)}-{revision_hash}",
        "sourceUrl": string_value(quote.get("driveSheetUrl") or quote.get("sourceUrl")),
        "sourceReceivedAt": now_iso(),
        "requesterName": string_value(quote.get("requesterName")),
        "requesterEmail": string_value(quote.get("requesterEmail")),
        "requesterPhone": string_value(quote.get("requesterPhone")),
        "agencyName": string_value(quote.get("agencyName")),
        "clientName": string_value(quote.get("clientName")),
        "audience": string_value(quote.get("audience")) or "OPERATOR",
        "language": string_value(quote.get("language")) or "en",
        "arrivalDate": date_only(quote.get("arrivalDate")),
        "departureDate": date_only(quote.get("departureDate")),
        "nights": coerce_int(quote.get("nights"), None),
        "guestCount": coerce_int(quote.get("guestCount"), None),
        "guideCount": coerce_int(quote.get("guideCount"), 0) or 0,
        "requestSummary": f"Revision of {quote_number}: {instruction}",
        "rawRequestText": f"Revision of {quote_number}: {instruction}",
        "requestedServices": quote.get("requestedServices") if isinstance(quote.get("requestedServices"), dict) else {},
        "calculation": calculation,
        "assumptions": assumptions,
        "missingFields": list_value(quote.get("missingFields")) or list_value(calculation.get("missingFields")),
        "sourceTrace": source_trace,
        "replyDraft": "Revised draft only; do not send.",
        "agentMetadata": {
            "agent": "Cotiza",
            "workflow": "quote-draft-revision",
            "revisionOfQuoteId": quote_id,
            "revisionOfQuoteNumber": quote_number,
            "revisionInstruction": instruction,
            "quoteRuleVersion": QUOTE_RULE_VERSION,
        },
        "idempotencyKey": f"revision-{normalize_id_part(quote_number)}-{revision_hash}-rules-{normalize_id_part(QUOTE_RULE_VERSION)}",
        "status": "DRAFTED",
    }
    return normalize_quote_create_payload(payload), revision_note


def create_revision_draft(payload: dict[str, Any], config: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any] | None, list[str]]:
    row = create_draft_row(payload, config)
    drive_result, warnings = create_drive_for_draft(row, payload)
    return row, drive_result, warnings


def mock_calculate(payload: dict[str, Any]) -> dict[str, Any]:
    summary = request_summary(payload).lower()
    year = rate_year_from_payload(payload)
    rates = quote_rates_for_year(year)
    guests_payload = object_value(payload, "guests")
    meals_payload = object_value(payload, "meals")
    birding_payload = object_value(payload, "birding")
    lodging_payload = object_value(payload, "lodging")
    audience = first_value(payload, "audience", "quoteAudience")
    if not isinstance(audience, str) or audience not in {"operator", "direct"}:
        if "operator" in summary or "agency" in summary:
            audience = "operator"
        elif "direct" in summary:
            audience = "direct"
        else:
            raise ToolError("missing_field", "Audience is required before pricing.")

    nights = coerce_int(first_value(payload, "nights", "nightCount"), None)
    if nights is None:
        nights = 2 if "two night" in summary or "2 night" in summary else 1

    guests = coerce_int(first_value(payload, "guestCount"), None)
    if guests is None:
        adults = coerce_int(guests_payload.get("adults"), 0) or 0
        children = coerce_int(guests_payload.get("children"), 0) or 0
        guests = adults + children if adults + children > 0 else None
    if guests is None:
        guests = 2 if "couple" in summary or "pareja" in summary else 1

    birding_days = coerce_int(first_value(payload, "birdingDays", "birdTourDays"), None)
    if birding_days is None:
        birding_days = coerce_int(birding_payload.get("morningTourDays"), None)
    if birding_days is None:
        birding_days = nights if "birding every day" in summary or "pajareo todos" in summary else 0

    meals_included = any(token in summary for token in ("food included", "meals included", "full board", "comida incluida", "alimentacion completa", "alimentación completa", "aalimentacion completa", "aalimentación completa"))
    lunches = coerce_int(first_value(payload, "lunchCount"), None)
    dinners = coerce_int(first_value(payload, "dinnerCount"), None)
    if lunches is None:
        lunches = coerce_int(meals_payload.get("lunches"), None)
    if dinners is None:
        dinners = coerce_int(meals_payload.get("dinners"), None)
    if meals_included:
        lunches = lunches if lunches is not None else default_cabin_lunch_count(nights, summary)
        dinners = dinners if dinners is not None else nights
    else:
        lunches = lunches or 0
        dinners = dinners or 0
    breakfasts = coerce_int(meals_payload.get("breakfasts"), None) or 0

    lodging_rate = rates["cabin_operator_net"] if audience == "operator" else rates["cabin_rack"]
    cabin_count = coerce_int(lodging_payload.get("cabinCount"), 1) or 1
    lodging_requested = lodging_payload.get("requested") is not False and cabin_count > 0
    line_items: list[dict[str, Any]] = []
    if lodging_requested:
        line_items.append({
            "description": "Cabin lodging",
            "notes": "Rate is for two people and includes breakfast.",
            "sourceRule": f"{year} cabin operator net" if audience == "operator" else f"{year} cabin rack",
            "unitPriceCop": lodging_rate,
            "quantity": nights * cabin_count,
            "totalCop": lodging_rate * nights * cabin_count,
        })
    guide_room_count = coerce_int(lodging_payload.get("guideRoomCount"), None)
    if guide_room_count is None:
        guide_room_count = coerce_int(lodging_payload.get("guide_room_count"), 0) or 0
    if lodging_requested and guide_room_count > 0:
        line_items.append({
            "serviceCode": "guide_room",
            "description": "Guide room",
            "notes": "Breakfast included.",
            "category": "lodging",
            "sourceRule": f"{year} guide room operator net",
            "unitPriceCop": rates["guide_room"],
            "quantity": nights * guide_room_count,
            "totalCop": rates["guide_room"] * nights * guide_room_count,
        })

    if lodging_requested and guests > 2 * cabin_count:
        qty = (guests - 2 * cabin_count) * nights
        line_items.append({
            "description": "Extra person",
            "notes": "Children under two are free.",
            "sourceRule": f"{year} extra person",
            "unitPriceCop": rates["extra_person"],
            "quantity": qty,
            "totalCop": rates["extra_person"] * qty,
        })

    if breakfasts:
        qty = guests * breakfasts
        line_items.append({
            "description": "Breakfast",
            "notes": "Client breakfast.",
            "sourceRule": f"{year} breakfast",
            "unitPriceCop": rates["breakfast"],
            "quantity": qty,
            "totalCop": rates["breakfast"] * qty,
        })

    if lunches:
        qty = guests * lunches
        line_items.append({
            "description": "Lunch",
            "notes": "Lunch is not included with lodging.",
            "sourceRule": f"{year} lunch",
            "unitPriceCop": rates["lunch"],
            "quantity": qty,
            "totalCop": rates["lunch"] * qty,
        })

    if dinners:
        qty = guests * dinners
        line_items.append({
            "description": "Dinner",
            "notes": "Dinner is not included with lodging.",
            "sourceRule": f"{year} dinner",
            "unitPriceCop": rates["dinner"],
            "quantity": qty,
            "totalCop": rates["dinner"] * qty,
        })

    if birding_days:
        qty = guests * birding_days
        bird_tour_rate = rates["bird_tour_rack"]
        bird_tour_source = f"{year} bird tour rack"
        if audience == "operator" and not lodging_requested and rates.get("bird_tour_operator_net"):
            bird_tour_rate = rates["bird_tour_operator_net"]
            bird_tour_source = f"{year} bird tour operator net"
        line_items.append({
            "description": "Bird Tour",
            "notes": "Morning birding per person.",
            "sourceRule": bird_tour_source,
            "unitPriceCop": bird_tour_rate,
            "quantity": qty,
            "totalCop": bird_tour_rate * qty,
        })

    total = sum(int(item["totalCop"]) for item in line_items)
    assumptions = [
        "Mock calculation only; production must use Operations /api/quotes/calculate.",
        "Cabin rate is for two people and includes breakfast.",
    ]
    if meals_included:
        assumptions.append("Food included was interpreted as lunch and dinner per night; breakfast is included with lodging.")
    if "transport" not in summary and "transporte" not in summary:
        assumptions.append("Internal 4x4 transport excluded.")
    return {
        "ok": True,
        "mock": True,
        "pricebookVersion": rates["pricebook_version"],
        "audience": audience,
        "currency": "COP",
        "lineItems": line_items,
        "subtotalCop": total,
        "discountCop": 0,
        "totalCop": total,
        "assumptions": assumptions,
        "missingFields": [],
        "flags": ["mock_pricing_not_for_production"],
    }


def mock_quote_id(payload: dict[str, Any]) -> tuple[str, str]:
    key = payload.get("idempotencyKey") or payload.get("idempotency_key") or json_dumps(payload)
    digest = hashlib.sha256(str(key).encode()).hexdigest()
    quote_id = f"mock-{digest[:18]}"
    quote_number = f"Q-2026-MOCK-{int(digest[:4], 16) % 10000:04d}"
    return quote_id, quote_number


def gmail_query(args: dict[str, Any]) -> str:
    parts: list[str] = []
    query = validate_text("query", args.get("query"), required=True, max_len=1000) or ""
    normalized_query, _query_source = normalize_gmail_search_query(query)
    if normalized_query:
        parts.append(normalized_query)
    sender = validate_text("from", args.get("from"), max_len=320)
    if sender:
        parts.append(f"from:{sender}")
    after = validate_text("after", args.get("after"), max_len=40)
    if after:
        parts.append(f"after:{after}")
    before = validate_text("before", args.get("before"), max_len=40)
    if before:
        parts.append(f"before:{before}")
    return " ".join(parts)


def gmail_label_id(service: Any, label_name: str | None) -> str | None:
    if not label_name:
        return None
    response = service.users().labels().list(userId="me").execute()
    for item in response.get("labels", []):
        if item.get("name") == label_name:
            return item.get("id")
    return None


def configured_gmail_quote_label(config: dict[str, Any]) -> str | None:
    label = cfg_env(config, "OWLSWATCH_GMAIL_QUOTE_LABEL")
    if not isinstance(label, str) or not label.strip() or label.strip() in {"*", "ALL", "all"}:
        return None
    return label.strip()


def require_gmail_quote_label_id(service: Any, label_name: str) -> str:
    label_id = gmail_label_id(service, label_name)
    if not label_id:
        raise ToolError("gmail_label_not_found", "Configured Gmail quote label was not found for the delegated account.")
    return label_id


def header_value(headers: list[dict[str, Any]], name: str) -> str:
    lowered = name.lower()
    for item in headers:
        if str(item.get("name", "")).lower() == lowered:
            value = item.get("value")
            return value if isinstance(value, str) else ""
    return ""


def gmail_source_url(thread_id: str) -> str:
    return f"https://mail.google.com/mail/u/0/#inbox/{urllib.parse.quote(thread_id)}"


def gmail_url_search_query(value: str) -> str | None:
    urls = re.findall(r"https?://mail\.google\.com/[^\s<>()]+", value or "", re.I)
    for url in urls:
        parsed = urllib.parse.urlparse(url)
        if parsed.netloc.lower() != "mail.google.com":
            continue
        fragment = urllib.parse.unquote_plus(parsed.fragment or "").strip()
        fragment_parts = [part for part in fragment.split("/") if part]
        if len(fragment_parts) >= 2 and fragment_parts[0] == "search":
            candidate = compact_text(fragment_parts[1])
            if candidate:
                return candidate
        query_params = urllib.parse.parse_qs(parsed.query)
        for key in ("q", "query", "search"):
            values = query_params.get(key)
            if values and compact_text(values[0]):
                return compact_text(values[0])
    remainder = re.sub(r"https?://mail\.google\.com/[^\s<>()]+", " ", value or "", flags=re.I)
    remainder = compact_text(remainder)
    return remainder or None


def gmail_url_message_token(value: str) -> str | None:
    urls = re.findall(r"https?://mail\.google\.com/[^\s<>()]+", value or "", re.I)
    for url in urls:
        parsed = urllib.parse.urlparse(url)
        if parsed.netloc.lower() != "mail.google.com":
            continue
        fragment = urllib.parse.unquote_plus(parsed.fragment or "").strip()
        fragment_parts = [part for part in fragment.split("/") if part]
        if len(fragment_parts) < 2 or fragment_parts[0] == "search":
            continue
        candidate = fragment_parts[-1].split("?")[0].strip()
        if re.match(r"^[A-Za-z0-9_-]{8,220}$", candidate):
            return candidate
    return None


def normalize_gmail_search_query(value: str) -> tuple[str, str]:
    text = compact_text(value)
    if "mail.google.com" not in text.lower():
        return text, "plain_text"
    query = gmail_url_search_query(text)
    if query:
        return query, "gmail_url_search"
    token = gmail_url_message_token(text)
    if token:
        return token, "gmail_url_message"
    return "", "gmail_url_unresolved"


def decode_gmail_body_data(value: str | None) -> str:
    if not value:
        return ""
    padding = "=" * (-len(value) % 4)
    try:
        return base64.urlsafe_b64decode((value + padding).encode()).decode("utf-8", errors="replace")
    except Exception:
        return ""


def strip_html(value: str) -> str:
    text = re.sub(r"(?is)<(script|style).*?>.*?</\\1>", " ", value)
    text = re.sub(r"(?s)<br\\s*/?>", "\n", text)
    text = re.sub(r"(?s)</p\\s*>", "\n", text)
    text = re.sub(r"(?s)<.*?>", " ", text)
    text = html.unescape(text)
    return re.sub(r"[ \\t\\r\\f\\v]+", " ", text).strip()


def extract_gmail_body(payload: dict[str, Any]) -> str:
    plain: list[str] = []
    html_parts: list[str] = []

    def walk(part: dict[str, Any]) -> None:
        mime_type = part.get("mimeType")
        body_data = (part.get("body") or {}).get("data")
        if mime_type == "text/plain" and body_data:
            plain.append(decode_gmail_body_data(body_data))
        elif mime_type == "text/html" and body_data:
            html_parts.append(strip_html(decode_gmail_body_data(body_data)))
        for child in part.get("parts") or []:
            if isinstance(child, dict):
                walk(child)

    walk(payload)
    text = "\n".join(item.strip() for item in plain if item.strip())
    if text:
        return text
    return "\n".join(item.strip() for item in html_parts if item.strip())


def gmail_thread_has_label(thread: dict[str, Any], label_id: str | None) -> bool:
    if not label_id:
        return True
    return any(label_id in (message.get("labelIds") or []) for message in thread.get("messages") or [])


def gmail_thread_to_match(thread: dict[str, Any]) -> dict[str, Any] | None:
    thread_id = thread.get("id")
    if not isinstance(thread_id, str):
        return None
    messages = thread.get("messages") or []
    message = messages[-1] if messages else {}
    headers = (message.get("payload") or {}).get("headers") or []
    return {
        "threadId": thread_id,
        "subject": header_value(headers, "Subject"),
        "from": header_value(headers, "From"),
        "date": header_value(headers, "Date"),
        "snippet": thread.get("snippet") or message.get("snippet") or "",
    }


def gmail_thread_metadata_match(service: Any, thread_id: str, label_id: str | None = None) -> dict[str, Any] | None:
    try:
        thread = service.users().threads().get(userId="me", id=thread_id, format="metadata", metadataHeaders=["Subject", "From", "Date"]).execute()
    except Exception:
        return None
    if not gmail_thread_has_label(thread, label_id):
        return None
    return gmail_thread_to_match(thread)


def resolve_gmail_browser_message_url(service: Any, token: str, label_id: str | None = None) -> dict[str, Any] | None:
    try:
        message = service.users().messages().get(userId="me", id=token, format="metadata").execute()
        thread_id = message.get("threadId")
        if isinstance(thread_id, str):
            match = gmail_thread_metadata_match(service, thread_id, label_id)
            if match:
                return match
    except Exception:
        pass

    match = gmail_thread_metadata_match(service, token, label_id)
    if match:
        return match

    for candidate in (token, f"rfc822msgid:{token}"):
        params: dict[str, Any] = {"userId": "me", "q": candidate, "maxResults": 1}
        if label_id:
            params["labelIds"] = [label_id]
        try:
            response = service.users().threads().list(**params).execute()
        except Exception:
            continue
        for item in response.get("threads", [])[:1]:
            thread_id = item.get("id")
            if isinstance(thread_id, str):
                match = gmail_thread_metadata_match(service, thread_id, label_id)
                if match:
                    return match
    return None


def tool_gmail_search_quote_threads(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"query", "from", "after", "before", "label", "maxResults"})
    config = load_config()
    query = validate_text("query", args.get("query"), required=True, max_len=1000) or ""
    normalized_query, query_source = normalize_gmail_search_query(query)
    sender = validate_text("from", args.get("from"), max_len=320)
    after = validate_text("after", args.get("after"), max_len=40)
    before = validate_text("before", args.get("before"), max_len=40)
    requested_label = validate_text("label", args.get("label"), max_len=200)
    max_results = coerce_int(args.get("maxResults"), 5) or 5
    if max_results < 1 or max_results > 10:
        raise ToolError("invalid_input", "maxResults must be between 1 and 10.")

    if query_source == "gmail_url_unresolved" and not any([sender, after, before]):
        return {
            "ok": True,
            "matches": [],
            "query": normalized_query,
            "querySource": query_source,
            "warning": "This Gmail browser URL does not include a searchable term. Ask for sender, subject, or date.",
        }

    if mocks_enabled(config):
        matches = []
        haystack = " ".join([normalized_query, sender or "", requested_label or ""]).lower()
        if any(term in haystack for term in ("linda", "jaguarundi", "burgess", "test")):
            matches.append({
                "threadId": "mock-thread-linda-jaguarundi",
                "subject": "Couple quote request",
                "from": "Linda <linda@example.com>",
                "date": "2026-05-05T15:30:00.000Z",
                "snippet": "A couple wants two nights, cabin, meals included, and birding every day.",
            })
        if "neptuno" in haystack or "trachsel" in haystack or normalized_query == "FMfcgzQfCDMHsnfFpCCpKBmgvzGXPRhM":
            matches.append({
                "threadId": "mock-thread-neptuno-trachsel",
                "subject": "Re: BLOQUEO DE SERVICIOS Trachsel x2 Booking No.: NPFI115855_Owl's Watch",
                "from": "Promotora Neptuno <reservas@example.com>",
                "date": "2026-05-06T13:42:00.000Z",
                "snippet": "Latest reply asks for the final total after accepting Dec 28-31, 2026.",
            })
        return {"ok": True, "mock": True, "matches": matches[:max_results], "query": normalized_query, "querySource": query_source}

    service = google_build_service(config, "gmail", "v1", ["https://www.googleapis.com/auth/gmail.readonly"], delegated_subject=gmail_account(config))
    label_name = configured_gmail_quote_label(config)
    if label_name and requested_label and requested_label != label_name:
        raise ToolError("invalid_input", "Gmail searches are restricted to the configured quote label.")
    label_id = None
    if label_name or requested_label:
        label_id = require_gmail_quote_label_id(service, label_name or requested_label or "")

    if query_source == "gmail_url_message":
        match = resolve_gmail_browser_message_url(service, normalized_query, label_id)
        if match:
            return {"ok": True, "matches": [match], "query": normalized_query, "querySource": query_source, "resolution": "gmail_url_message_resolved"}

    params: dict[str, Any] = {"userId": "me", "q": gmail_query(args), "maxResults": max_results}
    if label_id:
        params["labelIds"] = [label_id]
    response = service.users().threads().list(**params).execute()
    matches: list[dict[str, Any]] = []
    for item in response.get("threads", [])[:max_results]:
        thread_id = item.get("id")
        if not isinstance(thread_id, str):
            continue
        match = gmail_thread_metadata_match(service, thread_id, label_id)
        if match:
            matches.append(match)
    result: dict[str, Any] = {"ok": True, "matches": matches, "query": normalized_query, "querySource": query_source}
    if query_source == "gmail_url_message" and not matches:
        result["warning"] = "This Gmail inbox URL token could not be resolved by Gmail API. Ask for sender, subject, or date."
    return result


def tool_gmail_read_thread(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"threadId"})
    config = load_config()
    thread_id = validate_safe_id("threadId", args.get("threadId"))
    if mocks_enabled(config) and thread_id == "mock-thread-linda-jaguarundi":
        return {
            "ok": True,
            "mock": True,
            "thread": {
                "threadId": thread_id,
                "sourceUrl": "mock://gmail/thread/mock-thread-linda-jaguarundi",
                "messages": [
                    {
                        "from": "Linda <linda@example.com>",
                        "to": "info@owlswatch.com",
                        "date": "2026-05-05T15:30:00.000Z",
                        "subject": "Couple quote request",
                        "bodyText": "Jaguarundi has a couple asking for two nights in a cabin, meals included, and birding every day. Operator rate please.",
                    }
                ],
            },
        }
    if mocks_enabled(config) and thread_id == "mock-thread-neptuno-trachsel":
        return {
            "ok": True,
            "mock": True,
            "thread": {
                "threadId": thread_id,
                "sourceUrl": "mock://gmail/thread/mock-thread-neptuno-trachsel",
                "messages": [
                    {
                        "from": "Promotora Neptuno <reservas@example.com>",
                        "to": "info@owlswatch.com",
                        "date": "2026-05-04T09:20:00.000Z",
                        "subject": "BLOQUEO DE SERVICIOS Trachsel x2 Booking No.: NPFI115855_Owl's Watch",
                        "bodyText": "Operator Promotora Neptuno / Dreamtime Travel AG. Guest Trachsel x2. Request Dec 28 2026 - Jan 1 2027, double cabin, full board.",
                    },
                    {
                        "from": "Valeria <info@owlswatch.com>",
                        "to": "Promotora Neptuno <reservas@example.com>",
                        "date": "2026-05-04T13:15:00.000Z",
                        "subject": "Re: BLOQUEO DE SERVICIOS Trachsel x2 Booking No.: NPFI115855_Owl's Watch",
                        "bodyText": "We can offer availability Dec 28-31, 2026 only because Owl's Watch closes Dec 31 at noon.",
                    },
                    {
                        "from": "Promotora Neptuno <reservas@example.com>",
                        "to": "info@owlswatch.com",
                        "date": "2026-05-06T13:42:00.000Z",
                        "subject": "Re: BLOQUEO DE SERVICIOS Trachsel x2 Booking No.: NPFI115855_Owl's Watch",
                        "bodyText": "Thank you, accepted for Dec 28-31, 2026. Please send the final total for Trachsel x2.",
                    },
                ],
            },
        }
    if mocks_enabled(config):
        raise ToolError("not_found", "Mock Gmail thread was not found.")
    service = google_build_service(config, "gmail", "v1", ["https://www.googleapis.com/auth/gmail.readonly"], delegated_subject=gmail_account(config))
    thread = service.users().threads().get(userId="me", id=thread_id, format="full").execute()
    label_name = configured_gmail_quote_label(config)
    if label_name:
        label_id = require_gmail_quote_label_id(service, label_name)
        if not any(label_id in (message.get("labelIds") or []) for message in thread.get("messages") or []):
            raise ToolError("gmail_thread_outside_quote_label", "Gmail thread is outside the configured quote label.")
    messages: list[dict[str, Any]] = []
    for message in thread.get("messages") or []:
        payload = message.get("payload") or {}
        headers = payload.get("headers") or []
        messages.append({
            "from": header_value(headers, "From"),
            "to": header_value(headers, "To"),
            "date": header_value(headers, "Date"),
            "subject": header_value(headers, "Subject"),
            "bodyText": extract_gmail_body(payload),
        })
    return {"ok": True, "thread": {"threadId": thread_id, "sourceUrl": gmail_source_url(thread_id), "messages": messages}}


def tool_quote_calculate(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"payload"})
    config = load_config()
    raw_payload = validate_object("payload", args.get("payload"))
    payload = normalize_calculate_payload(raw_payload)
    policy_payload = {**raw_payload, **payload}
    if mocks_enabled(config):
        return apply_quote_display_policies(policy_payload, mock_calculate(payload))
    url = f"{operations_base_url(config)}/api/quotes/calculate"
    data = http_json("POST", url, payload, bearer_headers(config), timeout=45, attempts=1)
    if "ok" not in data:
        data["ok"] = True
    return apply_quote_display_policies(policy_payload, data)


def tool_quote_create_draft(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"payload", "prepared_quote", "source_metadata", "idempotency_key", "redo"})
    config = load_config()
    high_level = isinstance(args.get("prepared_quote"), dict)
    if high_level:
        prepared_quote = validate_object("prepared_quote", args.get("prepared_quote"))
        source_metadata = validate_object("source_metadata", args.get("source_metadata")) if isinstance(args.get("source_metadata"), dict) else {}
        idempotency_key = validate_text("idempotency_key", args.get("idempotency_key"), max_len=300)
        redo = redo_requested(prepared_quote, source_metadata, args.get("redo"))
        payload = normalize_quote_create_payload(create_payload_from_prepared(prepared_quote, source_metadata, idempotency_key, redo=redo))
    else:
        payload = normalize_quote_create_payload(validate_object("payload", args.get("payload")))
    missing = sheet_blocking_missing_details(payload)
    if missing:
        raise ToolError(
            "missing_required_quote_details",
            "Ask for missing quote details before creating an Operations draft row: " + ", ".join(missing) + ".",
        )
    row = create_draft_row(payload, config)
    if not high_level:
        return row

    drive_result, warnings = create_drive_for_draft(row, payload)
    total = money_value((payload.get("calculation") or {}).get("totalCop"))
    result = {
        **row,
        "totalCop": total,
        "redo": bool_value((payload.get("agentMetadata") or {}).get("redo")) if isinstance(payload.get("agentMetadata"), dict) else False,
        "assumptions": list_value(payload.get("assumptions")),
        "missingOptional": list_value(payload.get("missingFields")),
        "warnings": warnings,
    }
    if drive_result:
        result["drive"] = drive_result
        if drive_result.get("driveFileId"):
            result["driveFileId"] = drive_result.get("driveFileId")
        if drive_result.get("driveSheetUrl"):
            result["driveSheetUrl"] = drive_result.get("driveSheetUrl")
    try:
        quote_number = string_value(row.get("quoteNumber")) or "quote"
        agency = string_value(payload.get("agencyName")) or "operator"
        tool_cotiza_memory_log({
            "content": f"{quote_number} | {agency} | total COP {total:,} | status={row.get('status') or 'DRAFTED'} | {row.get('reviewUrl') or ''}".strip()
        })
    except Exception:
        pass
    return result


def tool_quote_update_drive(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"quoteId", "driveFileId", "driveSheetUrl"})
    config = load_config()
    quote_id = validate_safe_id("quoteId", args.get("quoteId"))
    drive_file_id = validate_safe_id("driveFileId", args.get("driveFileId"))
    drive_sheet_url = validate_text("driveSheetUrl", args.get("driveSheetUrl"), required=True, max_len=2000)
    payload = {"driveFileId": drive_file_id, "driveSheetUrl": drive_sheet_url}
    if mocks_enabled(config):
        path = MOCK_QUOTES_DIR / f"{quote_id}.json"
        if not path.exists():
            raise ToolError("not_found", "Mock quote row was not found.")
        record = json.loads(path.read_text())
        record["driveFileId"] = drive_file_id
        record["driveSheetUrl"] = drive_sheet_url
        record["updatedAt"] = now_iso()
        path.write_text(json.dumps(record, indent=2, ensure_ascii=False) + "\n")
        return {"ok": True, "mock": True, "quoteId": quote_id, "driveFileId": drive_file_id, "driveSheetUrl": drive_sheet_url}
    url = f"{operations_base_url(config)}/api/quotes/{urllib.parse.quote(quote_id)}/drive"
    data = http_json("PATCH", url, payload, bearer_headers(config), timeout=45, attempts=2)
    if "ok" not in data:
        data["ok"] = True
    return data


def tool_quote_revise_draft(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"quote_ref", "quoteNumber", "quoteId", "instruction", "source_metadata"})
    config = load_config()
    reference = args.get("quote_ref") or args.get("quoteNumber") or args.get("quoteId")
    instruction = validate_text("instruction", args.get("instruction"), required=True, max_len=2000) or ""
    source_metadata = validate_object("source_metadata", args.get("source_metadata")) if isinstance(args.get("source_metadata"), dict) else {}
    quote = fetch_quote_by_reference(quote_ref_text(reference), config)
    payload, revision_note = revision_payload_from_quote(quote, instruction, source_metadata)
    row, drive_result, warnings = create_revision_draft(payload, config)
    total = money_value((payload.get("calculation") or {}).get("totalCop"))
    result = {
        **row,
        "revisionOfQuoteId": string_value(quote.get("id") or quote.get("quoteId")),
        "revisionOfQuoteNumber": string_value(quote.get("quoteNumber")),
        "revisionInstruction": instruction,
        "revisionNote": revision_note,
        "totalCop": total,
        "assumptions": list_value(payload.get("assumptions")),
        "missingOptional": list_value(payload.get("missingFields")),
        "warnings": warnings,
    }
    if drive_result:
        result["drive"] = drive_result
        if drive_result.get("driveFileId"):
            result["driveFileId"] = drive_result.get("driveFileId")
        if drive_result.get("driveSheetUrl"):
            result["driveSheetUrl"] = drive_result.get("driveSheetUrl")
    try:
        tool_cotiza_memory_log({
            "content": f"{row.get('quoteNumber') or 'quote'} revised from {quote.get('quoteNumber') or quote.get('id')} | {revision_note} | total COP {total:,}".strip()
        })
    except Exception:
        pass
    return result


def safe_sheet_name(value: str | None) -> str:
    text = value or "Unknown"
    text = re.sub(r"[^A-Za-z0-9 ._+\-&'()]+", "", text).strip() or "Unknown"
    return text[:80]


MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def parse_iso_date(value: Any) -> dt.date | None:
    text = date_value(value)
    if not text or not DATE_RE.match(text):
        return None
    try:
        return dt.date.fromisoformat(text)
    except ValueError:
        return None


def formatted_date_range(arrival: Any, departure: Any) -> str:
    if not arrival and not departure:
        return ""
    start = parse_iso_date(arrival)
    end = parse_iso_date(departure)
    if start and end:
        if start == end:
            return f"{MONTHS[start.month - 1]} {start.day} {start.year}"
        if start.year == end.year and start.month == end.month:
            return f"{MONTHS[start.month - 1]} {start.day}-{end.day} {start.year}"
        if start.year == end.year:
            return f"{MONTHS[start.month - 1]} {start.day}-{MONTHS[end.month - 1]} {end.day} {start.year}"
        return f"{MONTHS[start.month - 1]} {start.day} {start.year}-{MONTHS[end.month - 1]} {end.day} {end.year}"
    return " - ".join(str(value or "") for value in (arrival, departure)).strip()


def quote_file_name(data: dict[str, Any]) -> str:
    agency = safe_sheet_name(data.get("agencyName") or "Operator")
    client_name = string_value(data.get("clientName"))
    client = safe_sheet_name(client_name) if client_name else ""
    dates = safe_sheet_name(formatted_date_range(data.get("arrivalDate"), data.get("departureDate")) or "Dates")
    quote_id = safe_sheet_name(validate_text("quoteNumber", data.get("quoteNumber"), required=True, max_len=80) or "ID")
    return " - ".join(part for part in [agency, client, dates, quote_id] if part)


def legacy_quote_file_name(data: dict[str, Any]) -> str:
    parts = [
        validate_text("quoteNumber", data.get("quoteNumber"), required=True, max_len=80) or "Quote",
        safe_sheet_name(data.get("agencyName")),
        safe_sheet_name(data.get("requesterName")),
    ]
    dates = " ".join(str(data.get(key) or "") for key in ("arrivalDate", "departureDate")).strip()
    if dates:
        parts.append(safe_sheet_name(dates))
    guests = coerce_int(data.get("guestCount"), None)
    if guests is not None:
        parts.append(f"{guests} guests")
    return " - ".join(part for part in parts if part)


def display_description(value: Any) -> str:
    text = string_value(value) or ""
    replacements = {
        "breakfast": "Client Breakfast",
        "lunch": "Client Lunch",
        "dinner": "Client Dinner",
        "photography and bird tour": "Bird Tour",
        "bird photography and tour": "Bird Tour",
        "guide/driver breakfast": "Guide/Driver Breakfast",
        "guide/driver lunch": "Guide/Driver Lunch (Discounted)",
        "guide/driver dinner": "Guide/Driver Dinner",
        "guide breakfast": "Guide Breakfast",
        "guide lunch": "Guide Lunch (Discounted)",
        "guide dinner": "Guide Dinner",
        "driver breakfast": "Driver Breakfast",
        "driver lunch": "Driver Lunch (Discounted)",
        "driver dinner": "Driver Dinner",
        "trip leader breakfast": "Trip Leader Breakfast",
        "trip leader lunch": "Trip Leader Lunch",
        "trip leader dinner": "Trip Leader Dinner",
    }
    return replacements.get(text.lower(), text)


def present_value(value: Any) -> Any:
    return "" if value is None else value


def money_value(value: Any) -> int:
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9.-]+", "", value)
        try:
            return int(round(float(cleaned)))
        except ValueError:
            return 0
    return 0


def quantity_value(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except ValueError:
            return 0
    return 0


def line_total(item: dict[str, Any]) -> int:
    total = money_value(item.get("totalCop") or item.get("amountCop"))
    if total:
        return total
    unit = money_value(item.get("unitPriceCop"))
    qty = quantity_value(item.get("quantity"))
    return int(round(unit * qty))


def line_text(item: dict[str, Any]) -> str:
    return " ".join(
        str(item.get(key) or "")
        for key in ("serviceCode", "description", "name", "category", "sourceRule", "notes")
    ).lower()


def item_identity_text(item: dict[str, Any]) -> str:
    return " ".join(
        str(item.get(key) or "")
        for key in ("serviceCode", "description", "name", "category")
    ).lower()


def is_guide_room_item(item: dict[str, Any]) -> bool:
    text = item_identity_text(item)
    return (
        "guide_room" in text
        or "guide room" in text
        or ("habitaci" in text and "gu" in text)
    )


def is_cabin_item(item: dict[str, Any]) -> bool:
    text = item_identity_text(item)
    return (
        "cabin" in text
        or "caba" in text
        or (item.get("category") == "lodging" and not is_guide_room_item(item))
    )


def is_overnight_lodging_item(item: dict[str, Any]) -> bool:
    return is_cabin_item(item) or is_guide_room_item(item)


def is_tour_item(item: dict[str, Any]) -> bool:
    text = line_text(item)
    return "tour" in text or "bird" in text or "aves" in text or item.get("category") == "activity"


def operator_context(payload: dict[str, Any], calc: dict[str, Any]) -> bool:
    if infer_audience(payload, request_summary(payload)) == "operator":
        return True
    version = string_value(calc.get("pricebookVersion")) or ""
    if "operator" in version.lower():
        return True
    assumptions = calc.get("assumptions")
    if isinstance(assumptions, list) and any("operator" in str(item).lower() for item in assumptions):
        return True
    return False


def copy_line_with_total(item: dict[str, Any], unit_price: int | None = None, total: int | None = None) -> dict[str, Any]:
    copied = dict(item)
    if unit_price is not None:
        copied["unitPriceCop"] = unit_price
    if total is not None:
        copied["totalCop"] = total
    return copied


def item_uses_operator_net_rate(item: dict[str, Any]) -> bool:
    text = line_text(item)
    return "operatornetratecop" in text.replace("_", "").replace(".", "").replace(" ", "") or "operator net" in text


def apply_operator_discount_policy(payload: dict[str, Any], calc: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(calc, dict) or not operator_context(payload, calc):
        return calc
    if money_value(calc.get("discountCop")):
        return calc

    raw_items = calc.get("lineItems") or calc.get("items") or []
    if not isinstance(raw_items, list):
        return calc
    items = [item for item in raw_items if isinstance(item, dict)]
    if not items:
        return calc

    cabin_indexes = [idx for idx, item in enumerate(items) if is_cabin_item(item)]
    discount_indexes = cabin_indexes or [idx for idx, item in enumerate(items) if is_tour_item(item)]
    if not discount_indexes:
        return calc

    display_items: list[dict[str, Any]] = []
    discount = 0
    subtotal = 0
    for idx, item in enumerate(items):
        original_total = line_total(item)
        unit = money_value(item.get("unitPriceCop"))
        qty = quantity_value(item.get("quantity")) or 1
        if idx in discount_indexes and (cabin_indexes or item_uses_operator_net_rate(item)):
            rack_unit = int(round(unit / 0.9)) if unit else 0
            rack_total = int(round(rack_unit * qty)) if rack_unit else int(round(original_total / 0.9))
            subtotal += rack_total
            discount += max(0, rack_total - original_total)
            display_items.append(copy_line_with_total(item, rack_unit or None, rack_total))
        else:
            subtotal += original_total
            display_items.append(copy_line_with_total(item, unit or None, original_total))

    if not cabin_indexes and discount == 0:
        tour_basis = sum(line_total(items[idx]) for idx in discount_indexes)
        discount = int(round(tour_basis * 0.10))

    total = subtotal - discount
    updated = dict(calc)
    updated["lineItems"] = display_items
    updated["subtotalCop"] = subtotal
    updated["discountCop"] = discount
    updated["totalCop"] = total
    assumptions = updated.get("assumptions")
    if isinstance(assumptions, list):
        note = "Operator discount: 10% off cabins; for tour-only quotes, 10% off the tour."
        if note not in assumptions:
            updated["assumptions"] = [*assumptions, note]
    return updated


def apply_staff_meal_policy(payload: dict[str, Any], calc: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(calc, dict):
        return calc
    raw_items = calc.get("lineItems") or calc.get("items") or []
    if not isinstance(raw_items, list):
        return calc
    items = [item for item in raw_items if isinstance(item, dict)]
    if any(str(item.get("category") or "") == "staff_meal" for item in items):
        return calc

    guide_count, driver_count = staff_role_counts(payload)
    staff_count = guide_count + driver_count
    if staff_count <= 0:
        return calc
    year = rate_year_from_payload(payload, calc)
    rates = quote_rates_for_year(year)
    label = staff_role_label(payload)
    service_prefix = "guide_driver" if guide_count > 0 and driver_count > 0 else ("driver" if driver_count > 0 else "guide")

    breakfast_days = staff_meal_day_count(payload, "breakfast", items)
    lunch_days = staff_meal_day_count(payload, "lunch", items)
    dinner_days = staff_meal_day_count(payload, "dinner", items)

    added_items: list[dict[str, Any]] = []
    if breakfast_days > 0:
        quantity = staff_count * breakfast_days
        added_items.append({
            "serviceCode": f"{service_prefix}_breakfast",
            "description": f"{label} Breakfast",
            "category": "staff_meal",
            "unit": "person_meal",
            "unitPriceCop": rates["guide_driver_breakfast"],
            "quantity": quantity,
            "sourceRule": f"{service_prefix}.breakfast.free",
            "notes": f"Breakfast for {label.lower()}s is complimentary.",
            "totalCop": 0,
        })

    lunch_rate = rates["guide_driver_lunch"]
    lunch_total = 0
    if lunch_days > 0:
        quantity = staff_count * lunch_days
        lunch_total = lunch_rate * quantity
        added_items.append({
            "serviceCode": f"{service_prefix}_lunch",
            "description": f"{label} Lunch",
            "category": "staff_meal",
            "unit": "person_meal",
            "unitPriceCop": lunch_rate,
            "quantity": quantity,
            "sourceRule": f"{service_prefix}.lunch.rateCop",
            "notes": f"Discounted {label.lower()} lunch.",
            "totalCop": lunch_total,
        })

    dinner_total = 0
    if dinner_days > 0:
        dinner_rate = rates["guide_driver_dinner"]
        quantity = staff_count * dinner_days
        dinner_total = dinner_rate * quantity
        added_items.append({
            "serviceCode": f"{service_prefix}_dinner",
            "description": f"{label} Dinner",
            "category": "staff_meal",
            "unit": "person_meal",
            "unitPriceCop": dinner_rate,
            "quantity": quantity,
            "sourceRule": f"{service_prefix}.dinner.rateCop",
            "notes": f"Discounted {label.lower()} dinner.",
            "totalCop": dinner_total,
        })
    if not added_items:
        return calc

    added_total = lunch_total + dinner_total
    updated_items = [
        *items,
        *added_items,
    ]
    updated = dict(calc)
    updated["lineItems"] = updated_items
    current_subtotal = money_value(calc.get("subtotalCop")) or sum(line_total(item) for item in items)
    current_total = money_value(calc.get("totalCop")) or (current_subtotal - money_value(calc.get("discountCop")))
    updated["subtotalCop"] = current_subtotal + added_total
    updated["totalCop"] = current_total + added_total
    assumptions = updated.get("assumptions")
    note_parts = []
    if breakfast_days > 0:
        note_parts.append(f"{label} breakfast is complimentary.")
    if lunch_days > 0:
        note_parts.append(f"{label} lunch is COP {lunch_rate:,}.")
    if dinner_total:
        note_parts.append(f"{label} dinner is COP {rates['guide_driver_dinner']:,}.")
    note = " ".join(note_parts)
    if isinstance(assumptions, list) and note not in assumptions:
        updated["assumptions"] = [*assumptions, note]
    return updated


def add_meal_line(items: list[dict[str, Any]], service_code: str, description: str, unit_price: int, quantity: int, source_rule: str, notes: str) -> dict[str, Any]:
    return {
        "serviceCode": service_code,
        "description": description,
        "category": "restaurant",
        "unit": "person_meal",
        "unitPriceCop": unit_price,
        "quantity": quantity,
        "sourceRule": source_rule,
        "notes": notes,
        "totalCop": unit_price * quantity,
    }


def apply_trip_leader_meal_policy(payload: dict[str, Any], calc: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(calc, dict):
        return calc
    raw_items = calc.get("lineItems") or calc.get("items") or []
    if not isinstance(raw_items, list):
        return calc
    items = [item for item in raw_items if isinstance(item, dict)]
    if any(str(item.get("serviceCode") or "").startswith("trip_leader_") for item in items):
        return calc

    trip_leaders = infer_trip_leader_count(payload)
    if trip_leaders <= 0:
        return calc
    summary = request_summary(payload)
    rates = quote_rates_for_year(rate_year_from_payload(payload, calc))
    days = infer_staff_meal_days(payload)
    quantity = trip_leaders * days
    additional: list[dict[str, Any]] = []
    if requests_breakfast(payload, summary):
        additional.append(add_meal_line(items, "trip_leader_breakfast", "Trip Leader Breakfast", rates["breakfast"], quantity, "restaurant.breakfast.operatorNetRateCop", "Trip leaders are billed as client-side participants for meals."))
    if requests_lunch(payload, summary):
        additional.append(add_meal_line(items, "trip_leader_lunch", "Trip Leader Lunch", rates["lunch"], quantity, "restaurant.lunch.operatorNetRateCop", "Trip leaders are billed as client-side participants for meals."))
    if requests_dinner(payload, summary):
        additional.append(add_meal_line(items, "trip_leader_dinner", "Trip Leader Dinner", rates["dinner"], quantity, "restaurant.dinner.operatorNetRateCop", "Trip leaders are billed as client-side participants for meals."))
    if not additional:
        return calc

    added_total = sum(line_total(item) for item in additional)
    updated = dict(calc)
    updated["lineItems"] = [*items, *additional]
    current_subtotal = money_value(calc.get("subtotalCop")) or sum(line_total(item) for item in items)
    current_total = money_value(calc.get("totalCop")) or (current_subtotal - money_value(calc.get("discountCop")))
    updated["subtotalCop"] = current_subtotal + added_total
    updated["totalCop"] = current_total + added_total
    assumptions = updated.get("assumptions")
    note = "Trip leaders are billed for requested meals but are not counted as bird-tour clients unless explicitly requested."
    if isinstance(assumptions, list) and note not in assumptions:
        updated["assumptions"] = [*assumptions, note]
    return updated


def apply_quote_display_policies(payload: dict[str, Any], calc: dict[str, Any]) -> dict[str, Any]:
    return apply_operator_discount_policy(payload, apply_trip_leader_meal_policy(payload, apply_staff_meal_policy(payload, calc)))


def image_formula(logo_url: str | None) -> str:
    if not logo_url:
        return ""
    escaped = logo_url.replace('"', '""')
    return f'=IMAGE("{escaped}",4,140,125)'


def quote_date_sequence(data: dict[str, Any]) -> list[dt.date]:
    start = parse_iso_date(date_only(data.get("arrivalDate")))
    end = parse_iso_date(date_only(data.get("departureDate")))
    if not start:
        return []
    if not end or end < start:
        end = start
    if (end - start).days > 31:
        return []
    return [start + dt.timedelta(days=offset) for offset in range((end - start).days + 1)]


def day_heading(value: dt.date) -> str:
    return f"{MONTHS[value.month - 1]} {value.day} {value.year}"


def is_day_header_row(row: list[Any]) -> bool:
    if not row or not isinstance(row[0], str):
        return False
    if row[0].startswith("::day::"):
        return True
    if any(cell not in ("", None) for cell in row[1:5]):
        return False
    return bool(re.match(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) \d{1,2} \d{4}$", row[0]))


def visible_day_header(row: list[Any]) -> list[Any]:
    if is_day_header_row(row):
        return [row[0].replace("::day::", "", 1), "", "", "", ""]
    return row


def meal_item_type(item: dict[str, Any]) -> str | None:
    text = line_text(item)
    if "breakfast" in text or "desayuno" in text:
        return "breakfast"
    if "lunch" in text or "almuerzo" in text:
        return "lunch"
    if "dinner" in text or "cena" in text or re.search(r"\bcomida\b", text):
        return "dinner"
    return None


def item_base_quantity(data: dict[str, Any], item: dict[str, Any]) -> int:
    text = line_text(item)
    if "guide_driver" in text or "guide/driver" in text or "staff_meal" in text:
        return max(1, infer_staff_count(data, "guide") + infer_staff_count(data, "driver"))
    if "trip_leader" in text:
        return max(1, infer_trip_leader_count(data))
    return max(1, coerce_int(data.get("guestCount"), None) or infer_guest_count(data, request_summary(data)) or 1)


def inferred_item_day_count(data: dict[str, Any], item: dict[str, Any]) -> int:
    quantity = quantity_value(item.get("quantity"))
    base = item_base_quantity(data, item)
    if quantity <= 0:
        return 1
    count = int(round(quantity / base)) if base else int(round(quantity))
    return max(1, count)


def day_indexes_for_item(data: dict[str, Any], dates: list[dt.date], item: dict[str, Any]) -> list[int]:
    if not dates:
        return []
    summary = request_summary(data)
    arrival = date_only(data.get("arrivalDate"))
    departure = date_only(data.get("departureDate"))
    nights = infer_nights(data, arrival, departure, summary) or 0
    day_trip = is_day_trip_quote(data, arrival, departure, nights, summary)
    day_count = min(len(dates), inferred_item_day_count(data, item))
    meal_type = meal_item_type(item)

    if is_overnight_lodging_item(item):
        overnight_count = nights or day_count
        return list(range(0, min(max(1, overnight_count), max(1, len(dates) - 1))))
    if day_trip:
        return [0]
    if meal_type == "dinner":
        return list(range(0, min(day_count, max(1, len(dates) - 1))))
    if meal_type == "lunch":
        if nights > 0 and day_count <= max(1, nights - 1) and len(dates) > 2:
            start = 1
        else:
            start = 0
        return list(range(start, min(len(dates), start + day_count)))
    if meal_type == "breakfast":
        if nights > 0 and not day_trip and money_value(item.get("unitPriceCop")) == 0:
            return list(range(1, min(len(dates), 1 + day_count)))
        return list(range(0, min(len(dates), day_count)))
    if is_tour_item(item):
        start = 1 if nights > 0 and len(dates) > 1 else 0
        return list(range(start, min(len(dates), start + day_count)))
    return [0]


def split_item_for_days(item: dict[str, Any], day_count: int) -> dict[str, Any]:
    copied = dict(item)
    if day_count <= 1:
        return copied
    quantity = quantity_value(item.get("quantity"))
    unit = money_value(item.get("unitPriceCop"))
    per_day_quantity = quantity / day_count if quantity else 0
    copied["quantity"] = int(per_day_quantity) if float(per_day_quantity).is_integer() else round(per_day_quantity, 2)
    copied["totalCop"] = int(round(unit * per_day_quantity)) if unit else int(round(line_total(item) / day_count))
    return copied


def included_breakfast_item(data: dict[str, Any]) -> dict[str, Any]:
    quantity = max(1, coerce_int(data.get("guestCount"), None) or infer_guest_count(data, request_summary(data)) or 1)
    return {
        "serviceCode": "included_lodging_breakfast",
        "description": "Client Breakfast",
        "category": "included_meal",
        "unit": "person_meal",
        "unitPriceCop": 0,
        "quantity": quantity,
        "sourceRule": "cabin.breakfast.included",
        "notes": "",
        "totalCop": 0,
    }


def included_breakfast_day_indexes(data: dict[str, Any], dates: list[dt.date], items: list[dict[str, Any]]) -> list[int]:
    if not dates or not any(is_cabin_item(item) for item in items):
        return []
    if any(is_client_included_breakfast_item(item) for item in items):
        return []
    summary = request_summary(data)
    arrival = date_only(data.get("arrivalDate"))
    departure = date_only(data.get("departureDate"))
    nights = infer_nights(data, arrival, departure, summary)
    if not nights or nights <= 0:
        return []
    return list(range(1, min(len(dates), nights + 1)))


def is_client_included_breakfast_item(item: dict[str, Any]) -> bool:
    if meal_item_type(item) != "breakfast" or money_value(item.get("unitPriceCop")) != 0:
        return False
    text = line_text(item)
    return (
        item.get("serviceCode") == "included_lodging_breakfast"
        or item.get("category") == "included_meal"
        or ("client" in text and "breakfast" in text)
    )


def quote_line_row(item: dict[str, Any]) -> list[Any]:
    return [
        display_description(item.get("description") or item.get("name")),
        display_notes(item),
        present_value(item.get("unitPriceCop")),
        present_value(item.get("quantity")),
        present_value(item.get("totalCop") if item.get("totalCop") is not None else item.get("amountCop")),
    ]


def display_notes(item: dict[str, Any]) -> str:
    notes = string_value(item.get("notes")) or ""
    text = line_text(item)
    if meal_item_type(item) == "breakfast" and money_value(item.get("unitPriceCop")) == 0:
        return ""
    if is_overnight_lodging_item(item) and re.search(r"\bbreakfast\b.{0,40}\bincluded\b|\bincluded\b.{0,40}\bbreakfast\b", notes, re.I):
        return ""
    if re.search(r"\bbreakfast\b.{0,40}\bcomplimentary\b|\bcomplimentary\b.{0,40}\bbreakfast\b", notes, re.I):
        return ""
    if "desayuno" in notes.lower() and ("inclu" in notes.lower() or "cortes" in notes.lower() or "gratis" in notes.lower()):
        return ""
    if "breakfast" in text and money_value(item.get("unitPriceCop")) == 0:
        return ""
    return notes


def day_display_sort_key(entry: tuple[int, dict[str, Any]]) -> tuple[int, int, int]:
    original_index, item = entry
    meal_type = meal_item_type(item)
    if is_cabin_item(item):
        return (0, 0, original_index)
    if is_guide_room_item(item):
        return (1, 0, original_index)
    if meal_type:
        meal_order = {"breakfast": 0, "lunch": 1, "dinner": 2}.get(meal_type, 9)
        return (2, meal_order, original_index)
    if is_tour_item(item):
        return (3, 0, original_index)
    return (4, 0, original_index)


def daily_quote_line_rows(data: dict[str, Any], line_items: list[Any]) -> list[list[Any]]:
    items = [item for item in line_items if isinstance(item, dict)]
    dates = quote_date_sequence(data)
    if not dates:
        return [quote_line_row(item) for item in items]

    grouped: dict[int, list[dict[str, Any]]] = {idx: [] for idx in range(len(dates))}
    for item in items:
        indexes = day_indexes_for_item(data, dates, item)
        if not indexes:
            grouped[0].append(item)
            continue
        for index in indexes:
            grouped[index].append(split_item_for_days(item, len(indexes)))
    for index in included_breakfast_day_indexes(data, dates, items):
        grouped[index].insert(0, included_breakfast_item(data))

    rows: list[list[Any]] = []
    for idx, date in enumerate(dates):
        day_items = grouped.get(idx) or []
        if not day_items:
            continue
        rows.append([f"::day::{day_heading(date)}", "", "", "", ""])
        rows.extend(quote_line_row(item) for _, item in sorted(enumerate(day_items), key=day_display_sort_key))
    return rows


def sheet_values(data: dict[str, Any], logo_url: str | None = None) -> dict[str, list[list[Any]]]:
    calc = validate_object("calculation", data.get("calculation"))
    calc = apply_quote_display_policies(data, calc)
    line_items = calc.get("lineItems") or calc.get("items") or []
    if not isinstance(line_items, list):
        line_items = []

    date_label = formatted_date_range(data.get("arrivalDate"), data.get("departureDate"))
    client = data.get("clientName") or ""
    operator = data.get("agencyName") or ""
    quote_rows: list[list[Any]] = [
        ["OWL'S WATCH", "", "", "", image_formula(logo_url)],
        ["NATURE RETREAT"],
        [],
        ["ID", data.get("quoteNumber") or ""],
        ["Date", date_label],
        ["Client", client],
        ["Operator", operator],
        [],
        ["Description", "Notes", "Price", "Quantity", "Total"],
    ]
    quote_rows.extend(visible_day_header(row) for row in daily_quote_line_rows(data, line_items))
    subtotal = calc.get("subtotalCop") or ""
    discount = calc.get("discountCop") or 0
    total = calc.get("totalCop") or calc.get("total") or ""
    quote_rows.append([])
    quote_rows.extend([["", "", "", "Subtotal", subtotal], ["", "", "", "Discount", discount]])
    quote_rows.append(["", "", "", "TOTAL", total])

    assumptions = validate_string_list("assumptions", data.get("assumptions"))
    missing = validate_string_list("missingFields", data.get("missingFields"))
    internal_rows = [
        ["Assumptions"],
        *[[item] for item in assumptions],
        [],
        ["Missing fields"],
        *[[item] for item in missing],
        [],
        ["Agent confidence", calc.get("confidence") or ""],
        ["Do not send until reviewed", "Yes"],
        [],
        ["Reply draft"],
        [data.get("replyDraft") or ""],
    ]

    source_trace = data.get("sourceTrace") if isinstance(data.get("sourceTrace"), dict) else {}
    trace_rows = [
        ["Source type", source_trace.get("sourceType") or ""],
        ["Source id", source_trace.get("threadId") or source_trace.get("messageId") or source_trace.get("id") or ""],
        ["Original summary", data.get("requestSummary") or ""],
        ["Pricebook version", calc.get("pricebookVersion") or ""],
        ["Rules used", ", ".join(str(item.get("sourceRule")) for item in line_items if isinstance(item, dict) and item.get("sourceRule"))],
    ]

    rule_rows = [["Description", "Source Rule", "Unit Price", "Notes"]]
    for item in line_items:
        if not isinstance(item, dict):
            continue
        rule_rows.append([
            item.get("description") or item.get("name") or "",
            item.get("sourceRule") or "",
            item.get("unitPriceCop") or "",
            item.get("notes") or "",
        ])

    return {
        "Quote": quote_rows,
        "Internal Notes": internal_rows,
        "Source Trace": trace_rows,
        "Price Rules": rule_rows,
    }


def rgb(hex_color: str) -> dict[str, float]:
    value = hex_color.lstrip("#")
    return {
        "red": int(value[0:2], 16) / 255,
        "green": int(value[2:4], 16) / 255,
        "blue": int(value[4:6], 16) / 255,
    }


def quote_row_counts(values: list[list[Any]]) -> tuple[int, int, int]:
    header_row = 8
    item_count = 0
    for row in values[9:]:
        if not any(cell != "" for cell in row):
            break
        item_count += 1
    total_row = len(values) - 1
    return header_row, item_count, total_row


def dimension_request(sheet_id: int, dimension: str, start: int, end: int, pixel_size: int) -> dict[str, Any]:
    return {
        "updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": dimension, "startIndex": start, "endIndex": end},
            "properties": {"pixelSize": pixel_size},
            "fields": "pixelSize",
        }
    }


def repeat_cell(sheet_id: int, start_row: int, end_row: int, start_col: int, end_col: int, cell: dict[str, Any], fields: str) -> dict[str, Any]:
    return {
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": start_row, "endRowIndex": end_row, "startColumnIndex": start_col, "endColumnIndex": end_col},
            "cell": cell,
            "fields": fields,
        }
    }


def border_request(sheet_id: int, start_row: int, end_row: int, start_col: int, end_col: int, style: str = "SOLID") -> dict[str, Any]:
    border = {"style": style, "width": 1, "color": rgb("#3F2A22")}
    return {
        "updateBorders": {
            "range": {"sheetId": sheet_id, "startRowIndex": start_row, "endRowIndex": end_row, "startColumnIndex": start_col, "endColumnIndex": end_col},
            "top": border,
            "bottom": border,
            "left": border,
            "right": border,
            "innerHorizontal": border,
            "innerVertical": border,
        }
    }


def quote_format_requests(sheet_ids: dict[str, int], values_by_tab: dict[str, list[list[Any]]]) -> list[dict[str, Any]]:
    quote_id = sheet_ids["Quote"]
    quote_values = values_by_tab["Quote"]
    header_row, item_count, total_row = quote_row_counts(quote_values)
    row_count = max(len(quote_values), total_row + 1)
    item_start = header_row + 1
    item_end = item_start + item_count
    item_end_safe = max(item_start, item_end)
    summary_start = max(item_end + 1, total_row - 2)
    day_rows = [idx for idx, row in enumerate(quote_values) if idx >= item_start and idx < item_end_safe and is_day_header_row(row)]
    requests: list[dict[str, Any]] = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": quote_id,
                    "gridProperties": {
                        "hideGridlines": False,
                        "frozenRowCount": 0,
                        "frozenColumnCount": 0,
                        "rowCount": row_count,
                        "columnCount": 5,
                    },
                },
                "fields": "gridProperties.hideGridlines,gridProperties.frozenRowCount,gridProperties.frozenColumnCount,gridProperties.rowCount,gridProperties.columnCount",
            }
        },
        {"unmergeCells": {"range": {"sheetId": quote_id, "startRowIndex": 0, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": 3}}},
        {"unmergeCells": {"range": {"sheetId": quote_id, "startRowIndex": item_start, "endRowIndex": max(item_start + 1, item_end_safe), "startColumnIndex": 0, "endColumnIndex": 5}}},
        {"mergeCells": {"range": {"sheetId": quote_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 3}, "mergeType": "MERGE_ALL"}},
        {"mergeCells": {"range": {"sheetId": quote_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": 3}, "mergeType": "MERGE_ALL"}},
        dimension_request(quote_id, "COLUMNS", 0, 1, 210),
        dimension_request(quote_id, "COLUMNS", 1, 2, 340),
        dimension_request(quote_id, "COLUMNS", 2, 3, 105),
        dimension_request(quote_id, "COLUMNS", 3, 4, 90),
        dimension_request(quote_id, "COLUMNS", 4, 5, 125),
        dimension_request(quote_id, "ROWS", 0, 1, 105),
        dimension_request(quote_id, "ROWS", header_row, header_row + 1, 34),
    ]
    requests.extend([
        repeat_cell(
            quote_id,
            0,
            1,
            0,
            3,
            {"userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 22, "foregroundColor": rgb("#3F2A22")}, "verticalAlignment": "MIDDLE"}},
            "userEnteredFormat(textFormat,verticalAlignment)",
        ),
        repeat_cell(
            quote_id,
            1,
            2,
            0,
            3,
            {"userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 11, "foregroundColor": rgb("#6D5449")}}},
            "userEnteredFormat(textFormat)",
        ),
        repeat_cell(
            quote_id,
            0,
            1,
            4,
            5,
            {"userEnteredFormat": {"horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"}},
            "userEnteredFormat(horizontalAlignment,verticalAlignment)",
        ),
        repeat_cell(
            quote_id,
            3,
            7,
            0,
            1,
            {"userEnteredFormat": {"backgroundColor": rgb("#E8D8C2"), "textFormat": {"bold": True, "foregroundColor": rgb("#3F2A22")}}},
            "userEnteredFormat(backgroundColor,textFormat)",
        ),
        repeat_cell(
            quote_id,
            3,
            7,
            1,
            3,
            {"userEnteredFormat": {"textFormat": {"fontSize": 12}, "wrapStrategy": "WRAP"}},
            "userEnteredFormat(textFormat,wrapStrategy)",
        ),
        repeat_cell(
            quote_id,
            header_row,
            header_row + 1,
            0,
            5,
            {"userEnteredFormat": {"backgroundColor": rgb("#3F2A22"), "textFormat": {"bold": True, "fontSize": 12, "foregroundColor": rgb("#FFFFFF")}, "horizontalAlignment": "CENTER"}},
            "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
        ),
        repeat_cell(
            quote_id,
            item_start,
            item_end_safe,
            0,
            5,
            {"userEnteredFormat": {"wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE", "textFormat": {"fontSize": 11}}},
            "userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)",
        ),
        repeat_cell(
            quote_id,
            item_start,
            item_end_safe,
            2,
            3,
            {"userEnteredFormat": {"numberFormat": {"type": "CURRENCY", "pattern": "$#,##0"}}},
            "userEnteredFormat.numberFormat",
        ),
        repeat_cell(
            quote_id,
            item_start,
            item_end_safe,
            4,
            5,
            {"userEnteredFormat": {"numberFormat": {"type": "CURRENCY", "pattern": "$#,##0"}}},
            "userEnteredFormat.numberFormat",
        ),
        repeat_cell(
            quote_id,
            item_start,
            item_end_safe,
            3,
            4,
            {"userEnteredFormat": {"horizontalAlignment": "RIGHT"}},
            "userEnteredFormat.horizontalAlignment",
        ),
        repeat_cell(
            quote_id,
            total_row,
            total_row + 1,
            3,
            5,
            {"userEnteredFormat": {"backgroundColor": rgb("#E8D8C2"), "textFormat": {"bold": True, "fontSize": 14, "foregroundColor": rgb("#111111")}, "numberFormat": {"type": "CURRENCY", "pattern": "$#,##0"}}},
            "userEnteredFormat(backgroundColor,textFormat,numberFormat)",
        ),
        repeat_cell(
            quote_id,
            summary_start,
            total_row,
            3,
            5,
            {"userEnteredFormat": {"textFormat": {"bold": True}, "numberFormat": {"type": "CURRENCY", "pattern": "$#,##0"}}},
            "userEnteredFormat(textFormat,numberFormat)",
        ),
        border_request(quote_id, header_row, max(header_row + 1, item_end), 0, 5),
        border_request(quote_id, summary_start, total_row + 1, 3, 5),
        border_request(quote_id, total_row, total_row + 1, 3, 5, "SOLID_THICK"),
    ])
    for row in day_rows:
        requests.extend([
            {"mergeCells": {"range": {"sheetId": quote_id, "startRowIndex": row, "endRowIndex": row + 1, "startColumnIndex": 0, "endColumnIndex": 5}, "mergeType": "MERGE_ALL"}},
            dimension_request(quote_id, "ROWS", row, row + 1, 28),
            repeat_cell(
                quote_id,
                row,
                row + 1,
                0,
                5,
                {"userEnteredFormat": {"backgroundColor": rgb("#D0D0D0"), "textFormat": {"bold": True, "fontSize": 12, "foregroundColor": rgb("#111111")}, "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"}},
                "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)",
            ),
        ])
    for title, sheet_id in sheet_ids.items():
        if title != "Quote":
            requests.append({"updateSheetProperties": {"properties": {"sheetId": sheet_id, "hidden": True}, "fields": "hidden"}})
    return requests


def drive_query_literal(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def ensure_quote_logo_url(drive: Any, folder_id: str) -> str | None:
    if not QUOTE_LOGO_PATH.is_file():
        return None
    name = "Owl's Watch Quote Logo.png"
    query = f"name='{drive_query_literal(name)}' and '{drive_query_literal(folder_id)}' in parents and trashed=false"
    logo_id: str | None = None
    try:
        existing = drive.files().list(
            q=query,
            spaces="drive",
            fields="files(id,name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files = existing.get("files") or []
        if files:
            logo_id = files[0].get("id")
        else:
            from googleapiclient.http import MediaFileUpload

            media = MediaFileUpload(str(QUOTE_LOGO_PATH), mimetype="image/png", resumable=False)
            uploaded = drive.files().create(
                body={"name": name, "parents": [folder_id]},
                media_body=media,
                fields="id",
                supportsAllDrives=True,
            ).execute()
            logo_id = uploaded.get("id")
        if logo_id:
            try:
                drive.permissions().create(
                    fileId=logo_id,
                    body={"type": "anyone", "role": "reader"},
                    fields="id",
                    supportsAllDrives=True,
                ).execute()
            except Exception:
                pass
            return f"https://drive.google.com/thumbnail?id={urllib.parse.quote(logo_id)}&sz=w320"
    except Exception:
        return None
    return None


def ensure_quote_tabs(sheets: Any, spreadsheet_id: str) -> dict[str, int]:
    metadata = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    titles = {item["properties"]["title"]: item["properties"]["sheetId"] for item in metadata.get("sheets", [])}
    requests: list[dict[str, Any]] = []
    if "Quote" not in titles:
        first = metadata["sheets"][0]["properties"]["sheetId"]
        requests.append({"updateSheetProperties": {"properties": {"sheetId": first, "title": "Quote"}, "fields": "title"}})
        titles["Quote"] = first
    for title in ("Internal Notes", "Source Trace", "Price Rules"):
        if title not in titles:
            requests.append({"addSheet": {"properties": {"title": title}}})
    if requests:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()
        metadata = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
        titles = {item["properties"]["title"]: item["properties"]["sheetId"] for item in metadata.get("sheets", [])}
    return {title: titles[title] for title in ("Quote", "Internal Notes", "Source Trace", "Price Rules")}


def write_quote_workbook(sheets: Any, spreadsheet_id: str, sheet_ids: dict[str, int], values_by_tab: dict[str, list[list[Any]]]) -> None:
    sheets.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id,
        body={"ranges": ["'Quote'!A1:Z200", "'Internal Notes'!A1:Z200", "'Source Trace'!A1:Z200", "'Price Rules'!A1:Z200"]},
    ).execute()
    value_ranges = [{"range": f"'{tab}'!A1", "values": values} for tab, values in values_by_tab.items()]
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "USER_ENTERED", "data": value_ranges},
    ).execute()
    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": quote_format_requests(sheet_ids, values_by_tab)},
    ).execute()


def missing_fields_text(data: dict[str, Any]) -> str:
    values = data.get("missingFields")
    if not isinstance(values, list):
        return ""
    return " | ".join(str(item).lower() for item in values)


def calculation_line_items(data: dict[str, Any]) -> list[dict[str, Any]]:
    calc = data.get("calculation")
    if not isinstance(calc, dict):
        return []
    items = calc.get("lineItems") or calc.get("items") or []
    return [item for item in items if isinstance(item, dict)] if isinstance(items, list) else []


def has_client_meal_line(data: dict[str, Any], meal: str) -> bool:
    for item in calculation_line_items(data):
        text = line_text(item)
        if "guide_driver" in text or "guide/driver" in text or "staff_meal" in text:
            continue
        if meal in text:
            return True
    return False


def sheet_has_visit_type(data: dict[str, Any]) -> bool:
    summary = request_summary(data).lower()
    if re.search(r"\b(cabin|caba[ñn]a|stay|nights?|lodging|hospedaje|birding|bird tour|aves|pajareo|day trip|pasad[ií]a)\b", summary, re.I):
        return True
    return any(is_cabin_item(item) or is_tour_item(item) for item in calculation_line_items(data))


def sheet_blocking_missing_details(data: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    missing_text = missing_fields_text(data)
    calc = data.get("calculation") if isinstance(data.get("calculation"), dict) else {}
    operator = operator_context(data, calc)

    def flagged(*terms: str) -> bool:
        return any(term in missing_text for term in terms)

    if not date_value(data.get("arrivalDate")):
        missing.append("dates/year")
    if data.get("guestCount") is None:
        missing.append("guest count")
    if not sheet_has_visit_type(data) or flagged("visit type", "cabin stay", "birding day trip"):
        missing.append("cabin stay or birding day trip")
    summary = request_summary(data)
    arrival = date_value(data.get("arrivalDate"))
    departure = date_value(data.get("departureDate"))
    nights = infer_nights(data, arrival, departure, summary)
    if is_day_trip_quote(data, arrival, departure, nights, summary):
        if requests_breakfast(data, summary) and not has_client_meal_line(data, "breakfast"):
            missing.append("client breakfast line item")
        if (requests_lunch(data, summary) or infer_birding_days(data, nights, summary) > 0) and not has_client_meal_line(data, "lunch"):
            missing.append("client lunch line item")
    return list(dict.fromkeys(missing))


def validate_sheet_ready(data: dict[str, Any]) -> None:
    missing = sheet_blocking_missing_details(data)
    if missing:
        raise ToolError(
            "missing_required_quote_details",
            "Ask for missing quote details before creating a Drive sheet: " + ", ".join(missing) + ".",
        )


def create_quote_xlsx(data: dict[str, Any], title: str) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        return None

    values_by_tab = sheet_values(data)
    wb = Workbook()
    default = wb.active
    default.title = "Quote"
    for title_name in ("Internal Notes", "Source Trace", "Price Rules"):
        wb.create_sheet(title_name)

    for tab, values in values_by_tab.items():
        ws = wb[tab]
        for row_idx, row in enumerate(values, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    quote = wb["Quote"]
    quote.merge_cells("A1:C1")
    quote.merge_cells("A2:C2")
    quote.column_dimensions["A"].width = 26
    quote.column_dimensions["B"].width = 40
    quote.column_dimensions["C"].width = 13
    quote.column_dimensions["D"].width = 11
    quote.column_dimensions["E"].width = 15
    quote.row_dimensions[1].height = 78
    quote.row_dimensions[9].height = 26
    quote.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    quote.page_setup.orientation = "landscape"
    quote.page_setup.paperSize = quote.PAPERSIZE_LETTER
    quote.page_setup.fitToWidth = 1
    quote.page_setup.fitToHeight = 1
    quote.page_margins.left = 0.25
    quote.page_margins.right = 0.25
    quote.page_margins.top = 0.25
    quote.page_margins.bottom = 0.25
    quote.print_area = f"A1:E{len(values_by_tab['Quote'])}"
    quote.sheet_view.showGridLines = False

    dark = "3F2A22"
    tan = "E8D8C2"
    light = "F7F2EA"
    thin = Side(style="thin", color=dark)
    thick = Side(style="medium", color=dark)
    quote["A1"].font = Font(bold=True, size=24, color=dark)
    quote["A1"].alignment = Alignment(vertical="center")
    quote["A2"].font = Font(bold=True, size=11, color="6D5449")

    for row in range(4, 8):
        quote.cell(row=row, column=1).font = Font(bold=True, color=dark)
        quote.cell(row=row, column=1).fill = PatternFill("solid", fgColor=tan)
        quote.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

    for cell in quote[9][:5]:
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    header_row, item_count, total_row = quote_row_counts(values_by_tab["Quote"])
    for row in range(10, 10 + item_count):
        for col in range(1, 6):
            cell = quote.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        quote.cell(row=row, column=3).number_format = '"$"#,##0'
        quote.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        quote.cell(row=row, column=5).number_format = '"$"#,##0'

    for row_idx, row_values in enumerate(values_by_tab["Quote"], 1):
        if row_idx >= 10 and is_day_header_row(row_values):
            quote.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            cell = quote.cell(row=row_idx, column=1)
            cell.fill = PatternFill("solid", fgColor="D0D0D0")
            cell.font = Font(bold=True, size=12, color="111111")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            quote.row_dimensions[row_idx].height = 22

    for col in range(4, 6):
        cell = quote.cell(row=total_row + 1, column=col)
        cell.fill = PatternFill("solid", fgColor=tan)
        cell.font = Font(bold=True, size=14)
        cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)
        if col == 5:
            cell.number_format = '"$"#,##0'

    for row in range(max(1, total_row - 1), total_row + 1):
        for col in range(4, 6):
            cell = quote.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if col == 5:
                cell.number_format = '"$"#,##0'

    if QUOTE_LOGO_PATH.is_file():
        logo = Image(str(QUOTE_LOGO_PATH))
        logo.width = 105
        logo.height = 118
        quote.add_image(logo, "E1")

    for tab in ("Internal Notes", "Source Trace", "Price Rules"):
        ws = wb[tab]
        ws.sheet_state = "hidden"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 80
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws["A1"].font = Font(bold=True, color=dark)
        ws["A1"].fill = PatternFill("solid", fgColor=light)

    out_dir = WORKSPACE / "spool" / "quote-sheets"
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_title = re.sub(r"[^A-Za-z0-9 ._+\-&'()]+", "", title).strip() or "quote"
    path = out_dir / f"{safe_title[:120]}.xlsx"
    wb.save(path)
    return path


def tool_drive_create_quote_sheet(args: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "quoteId",
        "quoteNumber",
        "agencyName",
        "requesterName",
        "clientName",
        "arrivalDate",
        "departureDate",
        "guestCount",
        "requestSummary",
        "calculation",
        "assumptions",
        "missingFields",
        "replyDraft",
        "sourceTrace",
    }
    validate_no_extra(args, allowed)
    config = load_config()
    quote_id = validate_safe_id("quoteId", args.get("quoteId"))
    validate_safe_id("quoteNumber", args.get("quoteNumber"))
    validate_object("calculation", args.get("calculation"))
    validate_sheet_ready(args)
    title = quote_file_name(args)

    if mocks_enabled(config):
        digest = hashlib.sha256(f"{quote_id}:{title}".encode()).hexdigest()
        drive_file_id = f"mock-drive-{digest[:16]}"
        MOCK_DRIVE_DIR.mkdir(parents=True, exist_ok=True)
        path = MOCK_DRIVE_DIR / f"{drive_file_id}.json"
        payload = {"title": title, "createdAt": now_iso(), "values": sheet_values(args)}
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
        return {
            "ok": True,
            "mock": True,
            "driveFileId": drive_file_id,
            "driveSheetUrl": f"mock://drive/{drive_file_id}",
            "localMockPath": str(path),
        }

    folder_id = google_folder_id(config)
    template_id = google_template_id(config)
    credentials_path = google_credentials_path(config)
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise ToolError("dependency_missing", "Google API Python packages are not installed in the tool runtime.") from exc

    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
    ]
    credentials = service_account.Credentials.from_service_account_file(credentials_path, scopes=scopes)
    sheets = build("sheets", "v4", credentials=credentials)
    drive = build("drive", "v3", credentials=credentials)

    try:
        used_xlsx = False
        if template_id:
            copied = drive.files().copy(
                fileId=template_id,
                body={"name": title, "parents": [folder_id]},
                fields="id,webViewLink",
                supportsAllDrives=True,
            ).execute()
            spreadsheet_id = copied["id"]
            spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
        else:
            xlsx_path = create_quote_xlsx(args, title)
            if xlsx_path:
                used_xlsx = True
                from googleapiclient.http import MediaFileUpload

                media = MediaFileUpload(
                    str(xlsx_path),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    resumable=False,
                )
                created = drive.files().create(
                    body={
                        "name": title,
                        "mimeType": "application/vnd.google-apps.spreadsheet",
                        "parents": [folder_id],
                    },
                    media_body=media,
                    fields="id,webViewLink",
                    supportsAllDrives=True,
                ).execute()
            else:
                created = drive.files().create(
                    body={
                        "name": title,
                        "mimeType": "application/vnd.google-apps.spreadsheet",
                        "parents": [folder_id],
                    },
                    fields="id,webViewLink",
                    supportsAllDrives=True,
                ).execute()
            spreadsheet_id = created["id"]
            spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
        if template_id or not used_xlsx:
            logo_url = ensure_quote_logo_url(drive, folder_id)
            sheet_ids = ensure_quote_tabs(sheets, spreadsheet_id)
            write_quote_workbook(sheets, spreadsheet_id, sheet_ids, sheet_values(args, logo_url))
    except Exception as exc:
        message = str(exc)
        if "sheets.googleapis.com" in message and "disabled" in message:
            raise ToolError("google_sheets_api_disabled", "Google Sheets API is not enabled for the configured service-account project.") from exc
        if "File not found" in message or "404" in message:
            raise ToolError("google_drive_folder_inaccessible", "Google Drive quote folder is not accessible to the configured service account.") from exc
        if "storage quota" in message.lower() or "storageQuotaExceeded" in message:
            raise ToolError(
                "google_drive_storage_quota_exceeded",
                "Google Drive refused to create the quote sheet because the configured service account has no Drive storage quota.",
            ) from exc
        raise ToolError("google_drive_error", "Google Drive quote sheet creation failed without exposing credential details.", retryable=True) from exc

    return {"ok": True, "driveFileId": spreadsheet_id, "driveSheetUrl": spreadsheet_url}


def tool_cotiza_memory_log(args: dict[str, Any]) -> dict[str, Any]:
    validate_no_extra(args, {"content"})
    content = validate_text("content", args.get("content"), required=True, max_len=4000)
    MEMORY_DIR.mkdir(parents=True, exist_ok=True)
    date = dt.datetime.now().strftime("%Y-%m-%d")
    path = MEMORY_DIR / f"{date}.md"
    with path.open("a") as f:
        f.write(f"- {local_now_label()} - {content.strip()}\n")
    return {"ok": True}


TOOLS: dict[str, tuple[str, dict[str, Any], Callable[[dict[str, Any]], dict[str, Any]]]] = {
    "owlswatch_gmail_search_quote_threads": ("Search read-only Owl's Watch Gmail quote threads by keywords or a Gmail browser URL.", {"type": "object", "properties": {"query": {"type": "string"}, "from": {"type": ["string", "null"]}, "after": {"type": ["string", "null"]}, "before": {"type": ["string", "null"]}, "label": {"type": ["string", "null"]}, "maxResults": {"type": "integer", "minimum": 1, "maximum": 10}}, "required": ["query"], "additionalProperties": False}, tool_gmail_search_quote_threads),
    "owlswatch_gmail_read_thread": ("Read one Owl's Watch Gmail quote thread by id.", {"type": "object", "properties": {"threadId": {"type": "string"}}, "required": ["threadId"], "additionalProperties": False}, tool_gmail_read_thread),
    "owlswatch_quote_prepare": ("Prepare a quote from raw request text, normalize defaults, validate missing info, and calculate a pricing preview.", {"type": "object", "properties": {"raw_text": {"type": "string"}, "source_metadata": {"type": "object", "additionalProperties": True}, "prior_context": {"type": ["object", "null"], "additionalProperties": True}, "user_overrides": {"type": ["object", "null"], "additionalProperties": True}, "parsed_intent": {"type": ["object", "null"], "additionalProperties": True}, "mode": {"type": ["string", "null"]}}, "required": ["raw_text"], "additionalProperties": False}, tool_quote_prepare),
    "owlswatch_quote_calculate": ("Calculate a quote using the Operations quote pricing endpoint.", {"type": "object", "properties": {"payload": {"type": "object", "additionalProperties": True}}, "required": ["payload"], "additionalProperties": False}, tool_quote_calculate),
    "owlswatch_quote_create_draft": ("Create an Operations quote draft and Drive sheet from a prepared quote. Set redo=true to create a fresh draft from an already-drafted source.", {"type": "object", "properties": {"prepared_quote": {"type": "object", "additionalProperties": True}, "source_metadata": {"type": "object", "additionalProperties": True}, "idempotency_key": {"type": ["string", "null"]}, "redo": {"type": ["boolean", "null"]}, "payload": {"type": "object", "additionalProperties": True}}, "additionalProperties": False}, tool_quote_create_draft),
    "owlswatch_quote_revise_draft": ("Create a revised draft/sheet from an existing draft quote and a simple instruction such as remove 2 lunches.", {"type": "object", "properties": {"quote_ref": {"type": "string"}, "quoteNumber": {"type": "string"}, "quoteId": {"type": "string"}, "instruction": {"type": "string"}, "source_metadata": {"type": "object", "additionalProperties": True}}, "required": ["instruction"], "additionalProperties": False}, tool_quote_revise_draft),
    "owlswatch_quote_update_drive": ("Patch an Operations quote row with its Google Drive draft link.", {"type": "object", "properties": {"quoteId": {"type": "string"}, "driveFileId": {"type": "string"}, "driveSheetUrl": {"type": "string"}}, "required": ["quoteId", "driveFileId", "driveSheetUrl"], "additionalProperties": False}, tool_quote_update_drive),
    "owlswatch_drive_create_quote_sheet": ("Create a Google Drive quote draft sheet in the configured folder.", {"type": "object", "properties": {"quoteId": {"type": "string"}, "quoteNumber": {"type": "string"}, "agencyName": {"type": ["string", "null"]}, "requesterName": {"type": ["string", "null"]}, "clientName": {"type": ["string", "null"]}, "arrivalDate": {"type": ["string", "null"]}, "departureDate": {"type": ["string", "null"]}, "guestCount": {"type": ["integer", "null"]}, "requestSummary": {"type": ["string", "null"]}, "calculation": {"type": "object", "additionalProperties": True}, "assumptions": {"type": "array", "items": {"type": "string"}}, "missingFields": {"type": "array", "items": {"type": "string"}}, "replyDraft": {"type": ["string", "null"]}, "sourceTrace": {"type": "object", "additionalProperties": True}}, "required": ["quoteId", "quoteNumber", "calculation"], "additionalProperties": False}, tool_drive_create_quote_sheet),
    "owlswatch_cotiza_memory_log": ("Append one quote summary line to Cotiza memory.", {"type": "object", "properties": {"content": {"type": "string"}}, "required": ["content"], "additionalProperties": False}, tool_cotiza_memory_log),
}


def rpc_result(request_id: Any, result: Any) -> dict[str, Any]:
    return {"jsonrpc": "2.0", "id": request_id, "result": result}


def rpc_error(request_id: Any, code: int, message: str) -> dict[str, Any]:
    return {"jsonrpc": "2.0", "id": request_id, "error": {"code": code, "message": message}}


def handle(request: dict[str, Any]) -> dict[str, Any] | None:
    method = request.get("method")
    request_id = request.get("id")
    if method == "initialize":
        return rpc_result(request_id, {"protocolVersion": "2024-11-05", "capabilities": {"tools": {}}, "serverInfo": {"name": "owlswatch_quotes", "version": "0.1.0"}})
    if method == "notifications/initialized":
        return None
    if method == "tools/list":
        return rpc_result(request_id, {"tools": [{"name": name, "description": desc, "inputSchema": schema} for name, (desc, schema, _) in TOOLS.items()]})
    if method == "tools/call":
        params = request.get("params") or {}
        name = params.get("name")
        args = params.get("arguments") or {}
        if name not in TOOLS or not isinstance(args, dict):
            return rpc_error(request_id, -32602, "Invalid tool call.")
        try:
            result = TOOLS[name][2](args)
        except Exception as exc:
            if os.environ.get("OWLSWATCH_DEBUG") == "1":
                traceback.print_exc(file=sys.stderr)
            result = sanitized_error(exc)
        return rpc_result(request_id, {"content": [{"type": "text", "text": json_dumps(result)}], "structuredContent": result, "isError": result.get("ok") is False})
    return rpc_error(request_id, -32601, "Method not found.")


def main() -> None:
    if len(sys.argv) == 3 and sys.argv[1] == "call":
        name = sys.argv[2]
        try:
            args = json.loads(sys.stdin.read() or "{}")
            if name not in TOOLS or not isinstance(args, dict):
                raise ToolError("invalid_tool", "Unknown tool or invalid arguments.")
            result = TOOLS[name][2](args)
        except Exception as exc:
            result = sanitized_error(exc)
        sys.stdout.write(json.dumps(result, ensure_ascii=False) + "\n")
        return
    for line in sys.stdin:
        if not line.strip():
            continue
        try:
            request = json.loads(line)
            response = handle(request)
        except Exception:
            response = rpc_error(None, -32700, "Parse or internal server error.")
        if response is not None:
            sys.stdout.write(json_dumps(response) + "\n")
            sys.stdout.flush()


if __name__ == "__main__":
    main()
